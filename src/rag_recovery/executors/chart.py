from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from pathlib import PurePosixPath

from lxml import etree
from openpyxl.utils.cell import range_boundaries

from ..models import Evidence, ExecutionResult, QueryPlan, Question
from ..normalize import nfkc, norm
from ..store import DocumentStore
from .base import Executor
from .utils import format_number, parse_number


@dataclass
class ChartInfo:
    source: str
    part: str
    title: str
    formulas: list[str]
    cached_series: list[list[float]]
    cached_categories: list[list[str]]


@dataclass
class EmbeddedChartExecutor(Executor):
    name: str = "chart"

    def execute(self, question: Question, plan: QueryPlan, store: DocumentStore) -> ExecutionResult:
        q = nfkc(question.text)
        project = plan.project_hints[0] if plan.project_hints else ""
        filename = plan.filename_hints[0] if plan.filename_hints else ""
        records = store.find(project_hint=project, filename_hint=filename, extensions={".xlsx", ".xlsm", ".pptx"}, selected_sources=question.selected_sources, limit=12)
        charts = [chart for rec in records for chart in self._extract_charts(rec)]
        if not charts:
            return ExecutionResult.abstain("埋め込みグラフ定義を取得できない")
        chart_no_match = re.search(r"グラフ\s*(\d+)", q)
        if chart_no_match:
            n = int(chart_no_match.group(1))
            charts = [c for c in charts if re.search(rf"chart{n}\.xml$", c.part)] or charts
        column_hits = self._column_names(charts, records, store)
        if "どのカラム" in q or "カラムを可視化" in q:
            if column_hits:
                return ExecutionResult(True, "、".join(dict.fromkeys(column_hits)), .95, "chart_source_column_resolution", [Evidence(c.source, c.part, f"title={c.title}, formulas={c.formulas}") for c in charts])
            return ExecutionResult.abstain("系列参照から列名を解決できない")
        if "最も多いカウント数" in q or ("ヒストグラム" in q and "カウント" in q):
            values = [value for chart in charts for series in chart.cached_series for value in series]
            if values:
                maximum = max(values)
                return ExecutionResult(True, format_number(maximum, q), .94, "chart_cached_max_count", [Evidence(c.source, c.part, f"cached_series={c.cached_series}") for c in charts])
            # Fallback: source range numeric frequencies; useful when chart cache is absent.
            frequency_max = self._frequency_from_source(charts, records, store)
            if frequency_max is not None:
                return ExecutionResult(True, str(frequency_max), .82, "chart_source_frequency_max", [Evidence(c.source, c.part, f"formulas={c.formulas}") for c in charts])
        if "最も低" in q or "最小" in q:
            pairs = []
            for chart in charts:
                for cats, vals in zip(chart.cached_categories, chart.cached_series):
                    if len(cats) == len(vals):
                        pairs.extend(zip(cats, vals))
            if pairs:
                cat, value = min(pairs, key=lambda x: x[1])
                return ExecutionResult(True, str(cat), .93, "chart_cached_argmin", [Evidence(c.source, c.part, f"pairs={pairs[:30]}") for c in charts])
        return ExecutionResult.abstain("グラフ質問をキャッシュ・参照範囲から解けない", diagnostics={"charts": [c.__dict__ for c in charts]})

    def _extract_charts(self, rec) -> list[ChartInfo]:
        charts = []
        try:
            with zipfile.ZipFile(rec.path) as zf:
                names = [n for n in zf.namelist() if re.search(r"/(?:charts)/chart\d+\.xml$", n)]
                for name in names:
                    root = etree.fromstring(zf.read(name))
                    title = " ".join(root.xpath("//*[local-name()='title']//*[local-name()='t']/text()"))
                    formulas = root.xpath("//*[local-name()='f']/text()")
                    cached_series = []
                    cached_categories = []
                    for ser in root.xpath("//*[local-name()='ser']"):
                        vals = []
                        for value in ser.xpath(".//*[local-name()='val']//*[local-name()='numCache']//*[local-name()='v']/text() | .//*[local-name()='yVal']//*[local-name()='numCache']//*[local-name()='v']/text()"):
                            num = parse_number(value)
                            if num is not None: vals.append(num)
                        if vals: cached_series.append(vals)
                        cats = ser.xpath(".//*[local-name()='cat']//*[local-name()='strCache']//*[local-name()='v']/text() | .//*[local-name()='xVal']//*[local-name()='strCache']//*[local-name()='v']/text()")
                        if not cats:
                            cats = ser.xpath(".//*[local-name()='cat']//*[local-name()='numCache']//*[local-name()='v']/text()")
                        if cats: cached_categories.append(list(map(str, cats)))
                    charts.append(ChartInfo(rec.relative_path, name, title, formulas, cached_series, cached_categories))
        except (zipfile.BadZipFile, etree.XMLSyntaxError):
            pass
        return charts

    def _column_names(self, charts, records, store):
        hits = []
        rec_map = {r.relative_path: r for r in records}
        for chart in charts:
            rec = rec_map.get(chart.source)
            if not rec or rec.extension not in {".xlsx", ".xlsm"}: continue
            wb = store.load_workbook(rec, data_only=True)
            for formula in chart.formulas:
                parsed = self._parse_ref(formula)
                if not parsed: continue
                sheet, cell_range = parsed
                if sheet not in wb.sheetnames: continue
                ws = wb[sheet]
                min_col, min_row, max_col, max_row = range_boundaries(cell_range)
                for col in range(min_col, max_col + 1):
                    # Series formulas often start below header; inspect row immediately above and first row.
                    for row in (max(1, min_row - 1), 1):
                        value = ws.cell(row, col).value
                        if value not in (None, "") and not isinstance(value, (int, float)):
                            hits.append(str(value))
                            break
        return hits

    def _frequency_from_source(self, charts, records, store):
        rec_map = {r.relative_path: r for r in records}
        values = []
        for chart in charts:
            rec = rec_map.get(chart.source)
            if not rec or rec.extension not in {".xlsx", ".xlsm"}: continue
            wb = store.load_workbook(rec, data_only=True)
            for formula in chart.formulas:
                parsed = self._parse_ref(formula)
                if not parsed: continue
                sheet, cell_range = parsed
                if sheet not in wb.sheetnames: continue
                ws = wb[sheet]
                for row in ws[cell_range]:
                    for cell in row:
                        num = parse_number(cell.value)
                        if num is not None: values.append(num)
        if not values: return None
        from collections import Counter
        return max(Counter(values).values())

    @staticmethod
    def _parse_ref(formula: str):
        value = formula.strip().lstrip("=")
        if "!" not in value: return None
        sheet, rng = value.rsplit("!", 1)
        sheet = sheet.strip("'").replace("''", "'")
        rng = rng.replace("$", "")
        if re.fullmatch(r"[A-Z]+\d+(?::[A-Z]+\d+)?", rng):
            return sheet, rng
        return None
