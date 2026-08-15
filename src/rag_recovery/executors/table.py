from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import date
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl.cell.cell import Cell

from ..models import Evidence, ExecutionResult, QueryPlan, Question
from ..normalize import nfkc, norm
from ..store import DocumentStore
from .base import Executor
from .utils import DATE_RE, apply_conditions, format_number, parse_date, parse_explicit_conditions, parse_number, read_table_file, resolve_column

ID_PATTERN = re.compile(r"\b(?:MS|CP|T|A|M)\d+\b", re.I)


def _rgb(cell: Cell) -> str:
    fill = cell.fill
    if fill.fill_type != "solid":
        return ""
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb[-6:].upper()
    if color.type == "indexed" and color.indexed is not None:
        return f"INDEXED:{color.indexed}"
    return ""


def _hsv_like(code: str) -> tuple[int, int, int] | None:
    if not code or code.startswith("INDEXED") or len(code) != 6:
        return None
    return tuple(int(code[i:i+2], 16) for i in (0, 2, 4))


def _is_color(code: str, name: str) -> bool:
    rgb = _hsv_like(code)
    if not rgb:
        return False
    r, g, b = rgb
    if name == "blue":
        return b >= 135 and b > r * 1.12 and b > g * 1.03
    if name == "yellow":
        return r >= 175 and g >= 145 and b <= 165 and abs(r - g) <= 100
    if name == "orange":
        return r >= 185 and 65 <= g <= 205 and b <= 175 and r > g
    if name == "red":
        return r >= 170 and r > g * 1.35 and r > b * 1.35
    return False


def _header_map(ws, max_rows: int = 30) -> tuple[int, dict[str, int]] | None:
    best = None
    for row in range(1, min(ws.max_row, max_rows) + 1):
        mapping = {norm(ws.cell(row, col).value): col for col in range(1, ws.max_column + 1) if ws.cell(row, col).value not in (None, "")}
        score = len(mapping)
        if best is None or score > best[0]:
            best = (score, row, mapping)
    return (best[1], best[2]) if best else None


def _find_col(mapping: dict[str, int], aliases: Iterable[str]) -> int | None:
    for alias in aliases:
        na = norm(alias)
        matches = [col for key, col in mapping.items() if na == key or na in key or key in na]
        if len(matches) == 1:
            return matches[0]
    return None


@dataclass
class TableQueryExecutor(Executor):
    name: str = "table"

    def execute(self, question: Question, plan: QueryPlan, store: DocumentStore) -> ExecutionResult:
        candidates = self._candidates(question, plan, store)
        if not candidates:
            return ExecutionResult.abstain("対象表計算ファイルを特定できない")
        handlers = (
            self._date_range_ids,
            self._style_based,
            self._autofilter,
            self._regression_prediction,
            self._id_count,
            self._person_task_count,
            self._generic_dataframe,
        )
        attempts = []
        for handler in handlers:
            result = handler(question, plan, store, candidates)
            if result is None:
                continue
            attempts.append({"handler": handler.__name__, "answered": result.answered, "reason": result.reason})
            if result.answered:
                result.diagnostics.setdefault("attempts", attempts)
                return result
        return ExecutionResult.abstain("表構造から一意回答に到達しない", diagnostics={"attempts": attempts, "candidates": [r.relative_path for r in candidates]})

    def _candidates(self, question, plan, store):
        project = plan.project_hints[0] if plan.project_hints else ""
        selected = question.selected_sources
        filename = plan.filename_hints[0] if plan.filename_hints else ""
        return store.find(project_hint=project, filename_hint=filename, extensions={".xlsx", ".xlsm", ".csv", ".tsv"}, selected_sources=selected, limit=12)

    def _date_range_ids(self, question, plan, store, candidates):
        q = nfkc(question.text)
        if not ("開始日" in q and "終了日" in q and ("タスクID" in q or "Task ID" in q)):
            return None
        dates = [date(*map(int, m)) for m in DATE_RE.findall(q)]
        if len(dates) < 2:
            return ExecutionResult.abstain("期間を解析できない")
        start, end = dates[:2]
        for rec in candidates:
            if rec.extension not in {".xlsx", ".xlsm"}:
                continue
            wb = store.load_workbook(rec, data_only=True)
            for ws in wb.worksheets:
                hm = _header_map(ws)
                if not hm:
                    continue
                hrow, mapping = hm
                id_col = _find_col(mapping, ["Task ID", "タスクID", "ID"])
                start_col = _find_col(mapping, ["開始日", "Start Date", "start"])
                end_col = _find_col(mapping, ["終了日", "End Date", "end"])
                if not (id_col and start_col and end_col):
                    continue
                hits = []
                for row in range(hrow + 1, ws.max_row + 1):
                    item = ws.cell(row, id_col).value
                    s, e = parse_date(ws.cell(row, start_col).value), parse_date(ws.cell(row, end_col).value)
                    if item not in (None, "") and ((s and start <= s <= end) or (e and start <= e <= end)):
                        hits.append((str(item), row, s, e))
                if hits:
                    return ExecutionResult(True, "、".join(x[0] for x in hits), .98, "table_date_range_filter", [Evidence(rec.relative_path, f"{ws.title}!row:{r}", f"id={i}, start={s}, end={e}") for i, r, s, e in hits])
        return ExecutionResult.abstain("ID・開始日・終了日列を特定できない")

    def _style_based(self, question, plan, store, candidates):
        q = nfkc(question.text)
        color = next((name for jp, name in (("青", "blue"), ("黄", "yellow"), ("オレンジ", "orange"), ("赤", "red")) if jp in q), "")
        if not color or not any(k in q for k in ("ハイライト", "色", "塗り")):
            return None
        for rec in candidates:
            if rec.extension not in {".xlsx", ".xlsm"}:
                continue
            wb_formula = store.load_workbook(rec, data_only=False)
            wb_value = store.load_workbook(rec, data_only=True)
            entries = []
            for ws in wb_formula.worksheets:
                wsv = wb_value[ws.title]
                hm = _header_map(ws)
                hrow, mapping = hm if hm else (1, {})
                for row in ws.iter_rows():
                    colored = [cell for cell in row if _is_color(_rgb(cell), color)]
                    if not colored:
                        continue
                    # `cell` is assigned only in the loop below.  Use the row
                    # coordinate here so a coloured row can be inspected before
                    # its individual cells are enumerated.
                    row_number = row[0].row
                    row_values = [wsv.cell(row_number, col).value for col in range(1, ws.max_column + 1)]
                    for cell in colored:
                        entries.append({"sheet": ws.title, "cell": cell.coordinate, "row": cell.row, "col": cell.column, "value": wsv[cell.coordinate].value, "row_values": row_values, "header": ws.cell(hrow, cell.column).value})
            if not entries:
                continue
            if "合計" in q:
                nums = [(e, parse_number(e["value"])) for e in entries]
                nums = [(e, v) for e, v in nums if v is not None]
                if nums:
                    total = sum(v for _, v in nums)
                    return ExecutionResult(True, format_number(total, q), .98, f"xlsx_{color}_fill_sum", [Evidence(rec.relative_path, f"{e['sheet']}!{e['cell']}", str(e["value"]), v) for e, v in nums], diagnostics={"raw_sum": total})
            if "行のタスク名" in q or "タスク名をすべて" in q:
                values = []
                for e in entries:
                    hm2 = _header_map(wb_formula[e["sheet"]])
                    if hm2:
                        _, mapping = hm2
                        col = _find_col(mapping, ["タスク名", "作業名", "タスク", "内容"])
                        if col:
                            value = wb_value[e["sheet"]].cell(e["row"], col).value
                            if value not in (None, ""):
                                values.append(str(value))
                values = list(dict.fromkeys(values))
                if values:
                    return ExecutionResult(True, "、".join(values), .96, f"xlsx_{color}_row_label_extract", [Evidence(rec.relative_path, f"{e['sheet']}!row:{e['row']}", str(e["row_values"])) for e in entries[:20]])
            # Yellow cells often annotate pivot/chart source condition and aggregate labels.
            if color == "yellow" and any(k in q for k in ("抽出条件", "集計内容", "対応するデータ")):
                # A request for conditions and aggregation needs the complete
                # highlighted pivot context.  A lone coloured cell elsewhere in
                # the workbook is not evidence for that context, so leave this
                # question to the specialised embedded-pivot executor.
                return ExecutionResult.abstain(
                    "yellow_pivot_requires_hierarchy_reconstruction",
                    diagnostics={"highlighted_cells": len(entries), "sheets": sorted({e["sheet"] for e in entries})},
                )
            if len(entries) <= 12:
                return ExecutionResult(True, "、".join(f"{e['cell']}={e['value']}" for e in entries), .82, f"xlsx_{color}_cell_extract", [Evidence(rec.relative_path, f"{e['sheet']}!{e['cell']}", str(e["value"])) for e in entries])
        return ExecutionResult.abstain(f"{color}色セルを検出できない")

    def _autofilter(self, question, plan, store, candidates):
        q = nfkc(question.text)
        if "フィルター" not in q or "条件" not in q:
            return None
        for rec in candidates:
            if rec.extension not in {".xlsx", ".xlsm"}:
                continue
            wb = store.load_workbook(rec, data_only=False)
            for ws in wb.worksheets:
                af = ws.auto_filter
                if not af or not af.ref:
                    continue
                ref = af.ref
                filters = []
                for fc in getattr(af, "filterColumn", []):
                    values = []
                    if fc.filters:
                        values.extend(getattr(fc.filters, "filter", []) or [])
                    if fc.customFilters:
                        values.extend(f"{x.operator}{x.val}" for x in fc.customFilters.customFilter)
                    filters.append((fc.colId, values))
                if filters:
                    return ExecutionResult(True, "、".join(f"列{col+1}={','.join(map(str, vals))}" for col, vals in filters), .96, "xlsx_autofilter_inspection", [Evidence(rec.relative_path, f"sheet:{ws.title}", f"ref={ref}, filters={filters}")])
        return ExecutionResult.abstain("AutoFilter条件を取得できない")

    def _regression_prediction(self, question, plan, store, candidates):
        q = nfkc(question.text)
        if not ("回帰" in q and "係数" in q and "予測値" in q):
            return None
        index_match = re.search(r"index\s*=\s*(\d+)", q, re.I)
        if not index_match:
            return ExecutionResult.abstain("対象indexを解析できない")
        target_index = int(index_match.group(1))
        for rec in candidates:
            if rec.extension not in {".xlsx", ".xlsm"}:
                continue
            wb_values = store.load_workbook(rec, data_only=True)
            coeffs: dict[str, float] = {}
            intercept: float | None = None
            target_row: dict[str, float] = {}
            for ws in wb_values.worksheets:
                # Coefficient tables: feature/name + coefficient/value.
                hm = _header_map(ws)
                if hm:
                    hrow, mapping = hm
                    feature_col = _find_col(mapping, ["feature", "変数", "項目", "説明変数"])
                    coef_col = _find_col(mapping, ["coefficient", "coef", "係数"])
                    index_col = _find_col(mapping, ["index", "id"])
                    if feature_col and coef_col:
                        for r in range(hrow + 1, ws.max_row + 1):
                            feature = ws.cell(r, feature_col).value
                            value = parse_number(ws.cell(r, coef_col).value)
                            if feature not in (None, "") and value is not None:
                                if norm(feature) in {"intercept", "切片", "const"}:
                                    intercept = value
                                else:
                                    coeffs[str(feature)] = value
                    if index_col:
                        headers = {col: ws.cell(hrow, col).value for col in range(1, ws.max_column + 1)}
                        for r in range(hrow + 1, ws.max_row + 1):
                            value = parse_number(ws.cell(r, index_col).value)
                            if value is not None and int(value) == target_index:
                                for col, header in headers.items():
                                    num = parse_number(ws.cell(r, col).value)
                                    if header not in (None, "") and num is not None:
                                        target_row[str(header)] = num
            common = [(feature, coef, target_row.get(feature)) for feature, coef in coeffs.items() if feature in target_row]
            if intercept is not None and common and len(common) >= max(1, len(coeffs) // 2):
                prediction = intercept + sum(coef * value for feature, coef, value in common if value is not None)
                evidence = [Evidence(rec.relative_path, "regression", f"intercept={intercept}", intercept)] + [Evidence(rec.relative_path, f"index={target_index}:{f}", f"coef={c}, value={v}, product={c*v}", c*v) for f, c, v in common if v is not None]
                return ExecutionResult(True, format_number(prediction, q), .96, "xlsx_regression_coefficient_application", evidence, diagnostics={"raw_prediction": prediction})
        return ExecutionResult.abstain("係数表と対象行を結合できない")

    def _id_count(self, question, plan, store, candidates):
        q = nfkc(question.text)
        requested = [prefix for jp, prefix in (("マイルストーンID", "MS"), ("タスクID", "T"), ("アクションID", "A"), ("チェックポイント", "CP")) if jp in q]
        if not requested or not any(k in q for k in ("合計", "いくつ", "何件")):
            return None
        found: dict[str, set[str]] = {p: set() for p in requested}
        evidence = []
        for rec in candidates:
            for unit in store.extract_text_units(rec):
                for item in ID_PATTERN.findall(unit.text):
                    upper = item.upper()
                    prefix = "MS" if upper.startswith("MS") else "CP" if upper.startswith("CP") else upper[0]
                    if prefix in found:
                        found[prefix].add(upper)
                if any(found.values()):
                    evidence.append(Evidence(rec.relative_path, unit.locator, ",".join(sorted(set(ID_PATTERN.findall(unit.text)), key=str))))
        total = sum(len(v) for v in found.values())
        if total:
            return ExecutionResult(True, str(total), .94, "identifier_nunique_count", evidence[:20], diagnostics={k: sorted(v) for k, v in found.items()})
        return ExecutionResult.abstain("指定IDを検出できない")

    def _person_task_count(self, question, plan, store, candidates):
        q = nfkc(question.text)
        m = re.search(r"([^、。\s]+さん)が担当者に含まれるタスクIDはいくつ", q)
        if not m:
            return None
        person = m.group(1).removesuffix("さん")
        for rec in candidates:
            for sheet, df in read_table_file(rec, store):
                id_col = next((resolve_column(df, x) for x in ("Task ID", "タスクID", "ID") if resolve_column(df, x)), None)
                person_col = next((resolve_column(df, x) for x in ("担当者", "Owner", "Assignee", "リソース") if resolve_column(df, x)), None)
                if id_col and person_col:
                    mask = df[person_col].astype(str).map(norm).str.contains(norm(person), regex=False)
                    ids = df.loc[mask, id_col].dropna().astype(str).unique().tolist()
                    if ids:
                        return ExecutionResult(True, str(len(ids)), .98, "table_person_task_nunique", [Evidence(rec.relative_path, f"sheet:{sheet}", f"person={person}, ids={ids}")])
        return ExecutionResult.abstain("担当者列とタスクID列を特定できない")

    def _generic_dataframe(self, question, plan, store, candidates):
        q = nfkc(question.text)
        for rec in candidates:
            for sheet, df in read_table_file(rec, store):
                if df.empty:
                    continue
                conditions = parse_explicit_conditions(q, df)
                filtered, notes = apply_conditions(df, conditions)
                # Special z-score conditions are parsed as a transformation, not a stored column.
                z_match = re.search(r"標準化された([A-Za-z_][A-Za-z0-9_]*)が\s*0未満", q)
                if z_match:
                    col = resolve_column(df, z_match.group(1))
                    if col:
                        s = pd.to_numeric(df[col], errors="coerce")
                        std = s.std(ddof=0)
                        if np.isfinite(std) and std > 0:
                            mask = ((s - s.mean()) / std) < 0
                            filtered = df.loc[mask].copy()
                            notes.append(f"zscore({col})<0: {int(mask.sum())} rows")
                target_col = self._target_column(q, df)
                group_col = self._group_column(q, df, target_col)
                if "割合" in q or "%" in q:
                    numerator_df = filtered
                    # 'X全体の平均を上回る' is evaluated against the referenced category population.
                    above = re.search(r"([A-Za-z_][A-Za-z0-9_]*)が([^、。]+?)全体の平均を上回る", q)
                    if above:
                        value_col = resolve_column(df, above.group(1))
                        category_expr = above.group(2)
                        category_conditions = parse_explicit_conditions(category_expr, df)
                        category_df, _ = apply_conditions(df, category_conditions)
                        if value_col and not category_df.empty:
                            threshold = pd.to_numeric(category_df[value_col], errors="coerce").mean()
                            numerator_df = numerator_df.loc[pd.to_numeric(numerator_df[value_col], errors="coerce") > threshold]
                            notes.append(f"{value_col}>{threshold}")
                    denom = len(filtered)
                    if denom:
                        value = len(numerator_df) / denom * 100
                        return ExecutionResult(True, format_number(value, q, unit="%"), .93, "table_filtered_ratio", [Evidence(rec.relative_path, f"sheet:{sheet}", f"numerator={len(numerator_df)}, denominator={denom}; {'; '.join(notes)}")])
                if target_col:
                    series = pd.to_numeric(filtered[target_col], errors="coerce")
                    if group_col and any(k in q for k in ("最も高", "最も低", "最も多", "上位")):
                        grouped = filtered.assign(_target=series).groupby(group_col, dropna=False)["_target"]
                        values = grouped.mean() if "平均" in q else grouped.sum() if any(k in q for k in ("合計", "総額")) else grouped.count()
                        values = values.dropna().sort_values(ascending=not any(k in q for k in ("最も高", "最も多", "上位")))
                        if not values.empty:
                            topn_match = re.search(r"上位\s*(\d+)", q)
                            topn = int(topn_match.group(1)) if topn_match else 1
                            selected = values.head(topn)
                            if topn == 1:
                                answer = str(selected.index[0])
                                if "総額とあわせて" in q or "金額で" in q:
                                    answer += f"（{format_number(float(selected.iloc[0]), q)}）"
                            else:
                                answer = "、".join(f"{idx}: {format_number(float(val), q)}" for idx, val in selected.items())
                            return ExecutionResult(True, answer, .94, "table_groupby_rank", [Evidence(rec.relative_path, f"sheet:{sheet}", selected.to_string())])
                    if "平均" in q and series.notna().any():
                        value = float(series.mean())
                        return ExecutionResult(True, format_number(value, q), .96, "table_filter_mean", [Evidence(rec.relative_path, f"sheet:{sheet}", f"rows={len(filtered)}, mean({target_col})={value}; {'; '.join(notes)}")])
                    if any(k in q for k in ("合計", "総額")) and series.notna().any():
                        value = float(series.sum())
                        return ExecutionResult(True, format_number(value, q), .96, "table_filter_sum", [Evidence(rec.relative_path, f"sheet:{sheet}", f"rows={len(filtered)}, sum({target_col})={value}; {'; '.join(notes)}")])
                if any(k in q for k in ("件数", "いくつ", "何行")) and (conditions or len(candidates) == 1):
                    return ExecutionResult(True, str(len(filtered)), .9, "table_filtered_count", [Evidence(rec.relative_path, f"sheet:{sheet}", f"rows={len(filtered)}; {'; '.join(notes)}")])
        return ExecutionResult.abstain("汎用表クエリを構築できない")

    @staticmethod
    def _target_column(q: str, df: pd.DataFrame) -> str | None:
        columns = sorted(map(str, df.columns), key=len, reverse=True)
        hits = [col for col in columns if norm(col) and norm(col) in norm(q)]
        # Exclude columns used only as explicit filters where possible.
        for col in hits:
            if any(word in q for word in (f"{col}の平均", f"{col}を", f"{col}が最も", f"{col}の合計", f"{col}は")):
                return col
        return hits[0] if hits else None

    @staticmethod
    def _group_column(q: str, df: pd.DataFrame, target: str | None) -> str | None:
        for col in sorted(map(str, df.columns), key=len, reverse=True):
            if col == target:
                continue
            if any(pattern in q for pattern in (f"{col}ごと", f"{col}別", f"{col}毎", f"最も高い{col}", f"{col}は")):
                return col
        return None
