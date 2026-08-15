import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook

from rag_competition.chart_executor import execute_chart_series_lookup
from rag_competition.schemas import FileRecord


def _record(path: Path) -> FileRecord:
    return FileRecord(
        file_id="chart", raw_path=str(path), relative_path=str(path), file_name=path.name,
        extension=".xlsx", size_bytes=0, modified_at="", sha1="x", area="", project_name="",
        major_folder="", document_kind="", version_label="",
    )


def _chart_workbook(path: Path, duplicate: bool = False, mismatch: bool = False) -> None:
    workbook = Workbook()
    workbook.active.title = "Sheet1"
    data = workbook.create_sheet("train")
    data["N1"] = "other" if mismatch else "hum"
    workbook.save(path)
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("xl/workbook.xml", '''<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/><sheet name="train" sheetId="2" r:id="rId2"/></sheets><definedNames><definedName name="_series">train!$N$1</definedName></definedNames></workbook>''')
        archive.writestr("xl/_rels/workbook.xml.rels", '''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="worksheet" Target="worksheets/sheet2.xml"/></Relationships>''')
        archive.writestr("xl/worksheets/sheet1.xml", '''<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><drawing r:id="rId1"/></worksheet>''')
        archive.writestr("xl/worksheets/_rels/sheet1.xml.rels", '''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="drawing" Target="../drawings/drawing1.xml"/></Relationships>''')
        count = 3 if duplicate else 2
        anchors = ''.join(f'''<xdr:twoCellAnchor><xdr:from><xdr:col>0</xdr:col><xdr:row>0</xdr:row></xdr:from><xdr:graphicFrame><xdr:nvGraphicFramePr><xdr:cNvPr name="\u30b0\u30e9\u30d5 1"/></xdr:nvGraphicFramePr><a:graphic><a:graphicData><cx:chart r:id="rId{index}"/></a:graphicData></a:graphic></xdr:graphicFrame></xdr:twoCellAnchor>''' for index in range(1, count))
        archive.writestr("xl/drawings/drawing1.xml", f'''<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:cx="http://schemas.microsoft.com/office/drawing/2014/chartex" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">{anchors}</xdr:wsDr>''')
        relations = ''.join(f'''<Relationship Id="rId{index}" Type="chartEx" Target="../charts/chartEx{index}.xml"/>''' for index in range(1, count))
        archive.writestr("xl/drawings/_rels/drawing1.xml.rels", f'''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{relations}</Relationships>''')
        for index in range(1, count):
            archive.writestr(f"xl/charts/chartEx{index}.xml", '''<cx:chartSpace xmlns:cx="http://schemas.microsoft.com/office/drawing/2014/chartex"><cx:chart><cx:plotArea><cx:plotAreaRegion><cx:series><cx:tx><cx:txData><cx:f>_series</cx:f><cx:v>hum</cx:v></cx:txData></cx:tx></cx:series></cx:plotAreaRegion></cx:plotArea></cx:chart></cx:chartSpace>''')


class ChartExecutorTests(unittest.TestCase):
    def test_native_chart_series_resolves_header_through_defined_name(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "chart.xlsx"
            _chart_workbook(path)
            output = execute_chart_series_lookup("Sheet1\u306e\u30b0\u30e9\u30d51\u306f\u3069\u306e\u30ab\u30e9\u30e0\u3092\u53ef\u8996\u5316\u3057\u307e\u3059\u304b", [_record(path)], Path("."))
        self.assertEqual("success", output["status"])
        self.assertEqual("hum", output["answer"])
        self.assertEqual("train!N1", output["evidence"][0]["source_range"])

    def test_duplicate_chart_number_is_suppressed(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "chart.xlsx"
            _chart_workbook(path, duplicate=True)
            output = execute_chart_series_lookup("Sheet1\u306e\u30b0\u30e9\u30d51\u306f\u3069\u306e\u30ab\u30e9\u30e0\u3092\u53ef\u8996\u5316\u3057\u307e\u3059\u304b", [_record(path)], Path("."))
        self.assertEqual("unsupported", output["status"])

    def test_series_header_mismatch_is_suppressed(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "chart.xlsx"
            _chart_workbook(path, mismatch=True)
            output = execute_chart_series_lookup("Sheet1\u306e\u30b0\u30e9\u30d51\u306f\u3069\u306e\u30ab\u30e9\u30e0\u3092\u53ef\u8996\u5316\u3057\u307e\u3059\u304b", [_record(path)], Path("."))
        self.assertEqual("unsupported", output["status"])

    def test_missing_native_chart_is_suppressed(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "chart.xlsx"
            workbook = Workbook()
            workbook.active.title = "Sheet1"
            workbook.save(path)
            output = execute_chart_series_lookup("Sheet1\u306e\u30b0\u30e9\u30d51\u306f\u3069\u306e\u30ab\u30e9\u30e0\u3092\u53ef\u8996\u5316\u3057\u307e\u3059\u304b", [_record(path)], Path("."))
        self.assertEqual("unsupported", output["status"])
