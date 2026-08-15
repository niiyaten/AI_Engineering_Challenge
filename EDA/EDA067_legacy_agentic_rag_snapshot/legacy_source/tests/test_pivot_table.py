from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from rag_competition.pivot_table import execute_pivot_extreme_question, extract_pivot_ir
from rag_competition.schemas import FileRecord
from rag_competition.source_requirements import infer_source_requirement, verify_selected_sources


def make_file(path: Path, project: str = "合成案件") -> FileRecord:
    return FileRecord(
        file_id="file_synthetic",
        raw_path=path.as_posix(),
        relative_path=path.name,
        file_name=path.name,
        extension=".xlsx",
        size_bytes=path.stat().st_size,
        modified_at="",
        sha1="synthetic",
        area="プロジェクト",
        project_name=project,
        major_folder="データ",
        document_kind="data",
        version_label="",
    )


def write_hierarchical_pivot(path: Path, tie: bool = False) -> FileRecord:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pivot"
    sheet.append(["行ラベル", "平均 / MonthlyIncome"])
    rows = [
        ("No", 5000, 0),
        ("Female", 6000, 1),
        ("Single", 7000, 2),
        ("Human Resources", 9000, 3),
        ("Technical Degree", 9000 if tie else 8000, 3),
        ("Male", 5500, 1),
        ("Married", 6500, 2),
        ("Life Sciences", 7500, 3),
        ("総計", 6100, 0),
    ]
    for label, value, indent in rows:
        sheet.append([label, value])
        sheet.cell(sheet.max_row, 1).alignment = sheet.cell(sheet.max_row, 1).alignment.copy(indent=indent)
    workbook.save(path)
    return make_file(path)


class PivotTableTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_reconstructs_row_hierarchy_and_totals(self) -> None:
        file = write_hierarchical_pivot(self.root / "train.xlsx")
        pivot = extract_pivot_ir(file, self.root / "train.xlsx")
        detail = [record for record in pivot.records if record.is_detail]
        self.assertEqual(pivot.row_axis_fields, ["level_1", "level_2", "level_3", "level_4"])
        self.assertEqual(
            detail[0].row_dimensions,
            {"level_1": "No", "level_2": "Female", "level_3": "Single", "level_4": "Human Resources"},
        )
        self.assertEqual(sum(record.is_grand_total for record in pivot.records), 1)
        self.assertEqual(sum(record.is_subtotal for record in pivot.records), 5)

    def test_extreme_answer_excludes_totals_and_is_reproducible(self) -> None:
        file = write_hierarchical_pivot(self.root / "train.xlsx")
        result = execute_pivot_extreme_question(
            1,
            "train.xlsxのPivotシートで平均月収が最も高い層の抽出条件を答えてください",
            file,
            self.root / "train.xlsx",
        )
        self.assertEqual(result["status"], "success")
        self.assertIn("Human Resources", result["answer"])
        self.assertTrue(result["verification"]["subtotal_handling_valid"])
        self.assertTrue(result["verification"]["reproducibility"])
        self.assertEqual(result["evidence"]["excluded_grand_total_count"], 1)

    def test_extreme_answer_suppresses_ambiguous_ties(self) -> None:
        file = write_hierarchical_pivot(self.root / "train.xlsx", tie=True)
        result = execute_pivot_extreme_question(
            1,
            "train.xlsxのPivotシートで平均月収が最も高い層の抽出条件を答えてください",
            file,
            self.root / "train.xlsx",
        )
        self.assertEqual(result["status"], "unsupported")
        self.assertEqual(result["failure_stage"], "uniqueness_failure")

    def test_source_requirement_requires_verified_relation(self) -> None:
        path = self.root / "train.xlsx"
        path.write_bytes(b"x")
        file = make_file(path, "別案件")
        requirement = infer_source_requirement("対象案件のtrain.xlsxを確認", required_projects=["対象案件"], required_file_types=["xlsx"])
        failed = verify_selected_sources(requirement, [file])
        passed = verify_selected_sources(requirement, [file], content_verified_file_ids={file.file_id})
        self.assertEqual(failed["verification_status"], "failed")
        self.assertEqual(passed["verification_status"], "passed")

    def test_version_pair_cardinality_is_checked(self) -> None:
        path = self.root / "v1.xlsx"
        path.write_bytes(b"x")
        requirement = infer_source_requirement("v1とv2の差分を比較してください")
        self.assertEqual(requirement.source_cardinality, "pair")
        result = verify_selected_sources(requirement, [make_file(path)], content_verified_file_ids={"file_synthetic"})
        self.assertFalse(result["source_cardinality_match"])


if __name__ == "__main__":
    unittest.main()
