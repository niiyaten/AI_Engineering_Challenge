from __future__ import annotations

import unittest
import tempfile
from types import SimpleNamespace
from pathlib import Path

from decimal import Decimal

from rag_competition.calculation_engine import (
    build_calculation_spec,
    calculate_ratio,
    execute_calculation_spec,
    linear_prediction,
    infer_coefficient_inputs,
    infer_coefficient_inputs_from_workbook,
    round_decimal,
    calculate_difference,
    rank_numeric_rows,
    calculate_schedule_effort,
    CalculationSpec,
    execute_operation_type,
    verify_operation_evidence,
)
from rag_competition.table_executor import table_filter


class CalculationEngineTest(unittest.TestCase):
    def setUp(self) -> None:
        rows = [
            {"id": "r1", "Age": "40", "EducationField": "Marketing", "MonthlyIncome": "12000"},
            {"id": "r2", "Age": "42", "EducationField": "Marketing", "MonthlyIncome": "13000"},
            {"id": "r3", "Age": "", "EducationField": "Marketing", "MonthlyIncome": "15000"},
            {"id": "r4", "Age": "60", "EducationField": "Sales", "MonthlyIncome": "14000"},
        ]
        self.table = SimpleNamespace(
            rows=rows,
            columns=["id", "Age", "EducationField", "MonthlyIncome"],
            file=SimpleNamespace(file_id="file_test", raw_path="raw/train.csv"),
            sheet_name="",
        )

    def test_multistep_mean_and_nearest_id(self) -> None:
        question = "train.csvでEducationFieldがMarketingかつMonthlyIncomeが10000より大きいデータを抽出し、Ageの平均値を計算し、その平均値に最も近い年齢のidをすべて答えてください。"
        spec = build_calculation_spec(question, self.table.columns)
        result = execute_calculation_spec(1, question, spec, self.table, table_filter)
        self.assertEqual("multi_step_calculation", spec.calculation_subtype)
        self.assertEqual("r1、r2", result["answer"])
        self.assertEqual("passed", result["verification"]["verification_status"])
        self.assertEqual(1, result["evidence"]["input_row_counts"]["missing_excluded"])

    def test_and_filters_are_structured(self) -> None:
        question = "EducationFieldがMarketingかつMonthlyIncomeが10000より大きいデータのAgeの平均値を計算してください。"
        spec = build_calculation_spec(question, self.table.columns)
        self.assertEqual(2, len(spec.filters))
        self.assertEqual("and", spec.logical_operator)

    def test_missing_values_are_not_zero(self) -> None:
        question = "EducationFieldがMarketingのデータのAgeの平均値を計算してください。"
        spec = build_calculation_spec(question, self.table.columns)
        result = execute_calculation_spec(1, question, spec, self.table, table_filter)
        self.assertEqual("41", result["answer"])
        self.assertEqual(1, result["evidence"]["input_row_counts"]["missing_excluded"])

    def test_ratio_and_percentage(self) -> None:
        ratio = calculate_ratio(1, 4, percentage=True, decimal_places=1)
        self.assertEqual("25.0", ratio["rounded_result"])
        self.assertEqual("0.25", ratio["raw_ratio"])

    def test_zero_denominator_is_suppressed(self) -> None:
        result = calculate_ratio(1, 0)
        self.assertEqual("zero_denominator", result["failure_stage"])

    def test_linear_prediction_aligns_by_name(self) -> None:
        result = linear_prediction("1.5", {"x2": "3", "x1": "2"}, {"x1": "0.5", "x2": "2"}, decimal_places=2)
        self.assertEqual("8.50", result["rounded_result"])
        self.assertEqual(["x1", "x2"], result["feature_names"])

    def test_linear_prediction_rejects_unaligned_names(self) -> None:
        result = linear_prediction(0, {"x1": 1}, {"x2": 1})
        self.assertEqual("coefficient_alignment_failure", result["failure_stage"])

    def test_coefficient_inputs_are_resolved_by_name_not_column_order(self) -> None:
        target = SimpleNamespace(sheet_name="data", columns=["id", "x2", "x1"], rows=[{"id": "7", "x2": "3", "x1": "2", "__row_number__": 2}])
        coefficients = SimpleNamespace(sheet_name="coefficients", columns=["feature", "coefficient"], rows=[
            {"feature": "x1", "coefficient": "0.5", "__row_number__": 2},
            {"feature": "x2", "coefficient": "2", "__row_number__": 3},
            {"feature": "intercept", "coefficient": "1.5", "__row_number__": 4},
        ])
        inputs = infer_coefficient_inputs([target, coefficients], "id=7の予測値")
        self.assertIsNotNone(inputs)
        result = linear_prediction(inputs["intercept"], inputs["feature_values"], inputs["coefficient_values"])
        self.assertEqual("8.5", result["rounded_result"])

    def test_workbook_coefficient_report_resolves_by_name_and_cell(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "model.xlsx"
            workbook = Workbook()
            report = workbook.active
            report.title = "Regression"
            report["B4"] = "coefficient"
            report["A5"] = "x2"; report["B5"] = "2"
            report["A6"] = "x1"; report["B6"] = "0.5"
            report["A7"] = "intercept"; report["B7"] = "1.5"
            data = workbook.create_sheet("Data")
            data.append(["id", "x1", "x2"])
            data.append(["7", "2", "3"])
            workbook.save(path)

            inputs = infer_coefficient_inputs_from_workbook(path, "id=7の予測値を小数第2位まで求めてください")

        self.assertIsNotNone(inputs)
        self.assertEqual({"x1", "x2"}, set(inputs["feature_values"]))
        self.assertEqual("Data!A2", inputs["key_cell"])
        self.assertIn("Regression!B5", inputs["cell_ranges"])
        result = linear_prediction(inputs["intercept"], inputs["feature_values"], inputs["coefficient_values"], decimal_places=2)
        self.assertEqual("8.50", result["rounded_result"])

    def test_workbook_prediction_suppresses_unbound_standardized_data(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "model.xlsx"
            workbook = Workbook()
            report = workbook.active
            report.title = "Regression"
            report["B4"] = "coefficient"
            report["A5"] = "x1"; report["B5"] = "1"
            report["A6"] = "intercept"; report["B6"] = "0"
            raw = workbook.create_sheet("Raw")
            raw.append(["id", "x1"]); raw.append(["7", "2"]); raw.append(["8", "4"])
            standardized = workbook.create_sheet("Prepared")
            standardized.append(["x1"])
            standardized["A2"] = "=STANDARDIZE(Raw!B2,AVERAGE(Raw!B$2:B$3),STDEV.P(Raw!B$2:B$3))"
            standardized["A3"] = "=STANDARDIZE(Raw!B3,AVERAGE(Raw!B$2:B$3),STDEV.P(Raw!B$2:B$3))"
            workbook.save(path)

            inputs = infer_coefficient_inputs_from_workbook(path, "id=7の予測値を求めてください")

        self.assertIsNone(inputs)

    def test_workbook_prediction_selects_formula_bound_standardized_representation(self) -> None:
        """A formula-linked standardized table wins only by coefficient re-production."""
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "model.xlsx"
            workbook = Workbook()
            report = workbook.active
            report.title = "Regression"
            report["A4"] = "variable"; report["B4"] = "coefficient"
            report["A5"] = "x1"; report["B5"] = "3"
            report["A6"] = "intercept"; report["B6"] = "2"
            raw = workbook.create_sheet("Raw")
            raw.append(["id", "x1", "y"])
            raw.append(["7", "1", "-1.6742346141747673"])
            raw.append(["8", "2", "2"])
            raw.append(["9", "3", "5.6742346141747673"])
            prepared = workbook.create_sheet("Prepared")
            prepared.append(["x1", "y"])
            for row in range(2, 5):
                prepared.cell(row, 1, f"=STANDARDIZE(Raw!B{row},AVERAGE(Raw!B$2:B$4),STDEV.P(Raw!B$2:B$4))")
                prepared.cell(row, 2, raw.cell(row, 3).value)
            workbook.save(path)

            inputs = infer_coefficient_inputs_from_workbook(path, "id=7の予測値を求めてください")

        self.assertIsNotNone(inputs)
        self.assertEqual("Prepared", inputs["selected_source_sheet"])
        self.assertEqual("STANDARDIZE", inputs["preprocessing"]["type"])
        self.assertLess(float(inputs["coefficient_reproduction_error"]), 0.000001)

    def test_workbook_prediction_suppresses_duplicate_target_rows(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "model.xlsx"
            workbook = Workbook()
            report = workbook.active
            report["B2"] = "coefficient"
            report["A3"] = "x1"; report["B3"] = "1"
            report["A4"] = "intercept"; report["B4"] = "0"
            data = workbook.create_sheet("Data")
            data.append(["id", "x1"]); data.append(["7", "2"]); data.append(["7", "3"])
            workbook.save(path)

            inputs = infer_coefficient_inputs_from_workbook(path, "id=7の予測値を求めてください")

        self.assertIsNone(inputs)

    def test_decimal_half_up(self) -> None:
        self.assertEqual(Decimal("2.35"), round_decimal(Decimal("2.345"), 2))

    def test_difference_supports_signed_and_absolute_results(self) -> None:
        self.assertEqual("3", calculate_difference("10", "7")["rounded_result"])
        self.assertEqual("3", calculate_difference("7", "10", absolute=True)["rounded_result"])

    def test_ranking_suppresses_unresolved_ties(self) -> None:
        rows = [{"name": "a", "value": "1"}, {"name": "b", "value": "1"}]
        result = rank_numeric_rows(rows, "value", return_column="name")
        self.assertEqual("uniqueness_failure", result["failure_stage"])

    def test_schedule_effort_requires_explicit_unit(self) -> None:
        result = calculate_schedule_effort("2", "3", unit="person_day")
        self.assertEqual("6", result["rounded_result"])
        self.assertEqual("schedule_spec_failure", calculate_schedule_effort("2", "3", unit="")["failure_stage"])

    def test_operation_type_dispatch_and_evidence(self) -> None:
        spec = CalculationSpec(operation_type="difference", absolute_or_signed="absolute")
        result = execute_operation_type(spec, {"left": "9", "right": "4"})
        verification = verify_operation_evidence(spec, {"left": "9", "right": "4"}, result)
        self.assertEqual("5", result["rounded_result"])
        self.assertEqual("passed", verification["verification_status"])


if __name__ == "__main__":
    unittest.main()
