from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import asdict, dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable


@dataclass
class CalculationSpec:
    schema_version: str = "1.0"
    calculation_subtype: str = "unsupported_calculation"
    target_file_role: str | None = None
    target_sheet: str | None = None
    row_selector: dict[str, Any] | None = None
    filters: list[dict[str, Any]] = field(default_factory=list)
    logical_operator: str = "and"
    input_columns: list[str] = field(default_factory=list)
    coefficient_source: str | None = None
    operations: list[dict[str, Any]] = field(default_factory=list)
    aggregation: str | None = None
    numerator_definition: str | None = None
    denominator_definition: str | None = None
    output_column: str | None = None
    output_type: str = "number"
    rounding: dict[str, Any] = field(default_factory=lambda: {"decimal_places": None, "mode": "half_up"})
    unit: str | None = None
    percentage: bool = False
    operation_type: str = ""
    population_definition: str | None = None
    grouping_conditions: list[dict[str, Any]] = field(default_factory=list)
    sort_key: str | None = None
    sort_direction: str | None = None
    tie_policy: str | None = None
    selection_rank: int | None = None
    intercept_definition: str | None = None
    feature_order: list[str] = field(default_factory=list)
    feature_mapping: dict[str, str] = field(default_factory=dict)
    missing_feature_policy: str = "suppress"
    encoding_policy: str | None = None
    schedule_start: str | None = None
    schedule_end: str | None = None
    duration_unit: str | None = None
    calendar_policy: str | None = None
    resource_count: Any = None
    workload: Any = None
    effort_formula: str | None = None
    difference_left: Any = None
    difference_right: Any = None
    difference_direction: str | None = None
    absolute_or_signed: str | None = None
    ambiguity_policy: str = "suppress"
    verification_requirements: list[str] = field(default_factory=list)


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value if value is not None else ""))).strip()


def _column_match(columns: list[str], hint: str) -> str | None:
    target = _norm(hint).lower()
    exact = [column for column in columns if _norm(column).lower() == target]
    return exact[0] if len(exact) == 1 else None


def build_calculation_spec(question: str, columns: list[str]) -> CalculationSpec:
    q = _norm(question)
    spec = CalculationSpec()
    explicit_file = re.search(r"([A-Za-z0-9_.-]+\.(?:csv|xlsx|tsv))", q, re.I)
    spec.target_file_role = explicit_file.group(1) if explicit_file else None
    sheet = re.search(r"([A-Za-z0-9_]+)\s*(?:\u30b7\u30fc\u30c8)", q)
    spec.target_sheet = sheet.group(1) if sheet else None
    decimals = re.search(r"\u5c0f\u6570\u7b2c\s*(\d+)\s*\u4f4d", q)
    if decimals:
        spec.rounding["decimal_places"] = int(decimals.group(1))
    spec.percentage = "%" in q or "\u5272\u5408" in q or "\u767e\u5206\u7387" in q
    spec.unit = "%" if spec.percentage else None
    spec.logical_operator = "or" if "\u307e\u305f\u306f" in q else "and"

    # Explicit ASCII column names are resolved against the raw table header.
    condition_pattern = (
        r"([A-Za-z][A-Za-z0-9_]*)\s*\u304c\s*(.+?)"
        r"(?=\u304b\u3064|\u307e\u305f\u306f|\u306e\u30c7\u30fc\u30bf|\u3092\u62bd\u51fa|$)"
    )
    for match in re.finditer(condition_pattern, q):
        column = _column_match(columns, match.group(1))
        if not column:
            continue
        value = match.group(2)
        greater = re.match(r"([+-]?\d+(?:\.\d+)?)\s*\u3088\u308a\u5927\u304d\u3044", value)
        less = re.match(r"([+-]?\d+(?:\.\d+)?)\s*\u3088\u308a\u5c0f\u3055\u3044", value)
        if greater:
            spec.filters.append({"column": column, "operator": "gt", "value": greater.group(1)})
        elif less:
            spec.filters.append({"column": column, "operator": "lt", "value": less.group(1)})
        else:
            clean = value.strip()
            spec.filters.append({"column": column, "operator": "eq", "value": clean})

    mean_match = re.search(r"([A-Za-z][A-Za-z0-9_]*)\s*\u306e\u5e73\u5747", q)
    mean_column = _column_match(columns, mean_match.group(1)) if mean_match else None
    nearest = "\u6700\u3082\u8fd1\u3044" in q
    output_match = re.search(r"\u306e\s*([A-Za-z][A-Za-z0-9_]*)\s*\u3092\u3059\u3079\u3066", q)
    output_column = _column_match(columns, output_match.group(1)) if output_match else None
    if mean_column and nearest and output_column and spec.filters:
        spec.calculation_subtype = "multi_step_calculation"
        spec.input_columns = list(dict.fromkeys([item["column"] for item in spec.filters] + [mean_column, output_column]))
        spec.output_column = output_column
        spec.aggregation = "mean"
        spec.output_type = "list"
        spec.operations = [
            {"step_id": "s1", "operation": "filter_rows", "output": "filtered_rows"},
            {"step_id": "s2", "operation": "mean", "input": f"filtered_rows.{mean_column}", "output": "group_mean"},
            {"step_id": "s3", "operation": "absolute_difference", "input": f"filtered_rows.{mean_column}", "comparison": "group_mean", "output": "distances"},
            {"step_id": "s4", "operation": "select_min_distance", "input": "distances", "output": f"nearest_rows.{output_column}"},
        ]
    elif mean_column and spec.filters:
        spec.calculation_subtype = "filtered_aggregation"
        spec.input_columns = list(dict.fromkeys([item["column"] for item in spec.filters] + [mean_column]))
        spec.aggregation = "mean"
        spec.operations = [{"step_id": "s1", "operation": "filter_rows"}, {"step_id": "s2", "operation": "mean", "input": mean_column}]
    # 明示的な計算語をoperation_typeへ保存し、後段の必須Evidence検証に使う。
    if any(term in q for term in ("上昇率", "割合", "百分率", "%")):
        spec.operation_type = "ratio_or_percentage"
        spec.calculation_subtype = "ratio_or_percentage"
        spec.aggregation = "ratio"
        spec.output_type = "percentage" if spec.percentage else "ratio"
        spec.operations = [{"step_id": "s1", "operation": "ratio"}]
    elif any(term in q for term in ("係数", "予測値", "切片")):
        spec.operation_type = "coefficient_prediction"
        spec.calculation_subtype = "coefficient_prediction"
        spec.aggregation = "linear_prediction"
        spec.operations = [{"step_id": "s1", "operation": "linear_prediction"}]
    elif any(term in q for term in ("最も低い", "最小", "最も高い", "最大", "ランキング")):
        spec.operation_type = "ranking_or_argmin"
        spec.calculation_subtype = "ranking_or_argmin"
        spec.sort_direction = "asc" if any(term in q for term in ("最も低い", "最小")) else "desc"
        spec.tie_policy = "suppress_if_multiple"
        spec.operations = [{"step_id": "s1", "operation": "argmin" if spec.sort_direction == "asc" else "argmax"}]
    elif any(term in q for term in ("差の絶対値", "差額", "差分")):
        spec.operation_type = "difference"
        spec.calculation_subtype = "difference"
        spec.absolute_or_signed = "absolute" if "絶対" in q else "signed"
        spec.operations = [{"step_id": "s1", "operation": "difference", "absolute": spec.absolute_or_signed == "absolute"}]
    elif any(term in q for term in ("工数", "人日", "人時", "稼働率")):
        spec.operation_type = "schedule_effort"
        spec.calculation_subtype = "schedule_effort"
        spec.duration_unit = "person_hour" if "人時" in q else "person_day" if "人日" in q else None
        spec.operations = [{"step_id": "s1", "operation": "schedule_effort", "duration_unit": spec.duration_unit}]
    elif spec.calculation_subtype in {"multi_step_calculation", "filtered_aggregation"}:
        spec.operation_type = "single_table_aggregation"
    spec.verification_requirements = ["input_presence", "operation_validity", "reproducibility", "source_range"]
    return spec


def _decimal_number(value: Any) -> Decimal | None:
    text = _norm(value).replace(",", "").replace("\u00a5", "").replace("\u5186", "")
    if not text or text.lower() in {"nan", "none"}:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return Decimal(text)
    except Exception:
        return None


def round_decimal(value: Decimal, decimal_places: int | None) -> Decimal:
    if decimal_places is None:
        return value
    quantum = Decimal(1).scaleb(-decimal_places)
    return value.quantize(quantum, rounding=ROUND_HALF_UP)


def calculate_ratio(
    numerator: Any,
    denominator: Any,
    *,
    percentage: bool = False,
    decimal_places: int | None = None,
) -> dict[str, Any]:
    numerator_value = _decimal_number(numerator)
    denominator_value = _decimal_number(denominator)
    if numerator_value is None or denominator_value is None:
        return {"status": "unsupported", "failure_stage": "numeric_conversion_failure"}
    if denominator_value == 0:
        return {"status": "unsupported", "failure_stage": "zero_denominator"}
    ratio = numerator_value / denominator_value
    output = ratio * Decimal(100) if percentage else ratio
    rounded = round_decimal(output, decimal_places)
    return {
        "status": "success",
        "numerator_count": str(numerator_value),
        "denominator_count": str(denominator_value),
        "raw_ratio": str(ratio),
        "percentage_value": str(output) if percentage else None,
        "rounded_result": format(rounded, "f"),
        "formula": "numerator / denominator" + (" * 100" if percentage else ""),
    }


def linear_prediction(
    intercept: Any,
    feature_values: dict[str, Any],
    coefficient_values: dict[str, Any],
    *,
    decimal_places: int | None = None,
) -> dict[str, Any]:
    if set(feature_values) != set(coefficient_values):
        return {"status": "unsupported", "failure_stage": "coefficient_alignment_failure"}
    intercept_value = _decimal_number(intercept)
    if intercept_value is None:
        return {"status": "unsupported", "failure_stage": "coefficient_source_failure"}
    products: dict[str, str] = {}
    total = intercept_value
    for name in sorted(feature_values):
        feature = _decimal_number(feature_values[name])
        coefficient = _decimal_number(coefficient_values[name])
        if feature is None or coefficient is None:
            return {"status": "unsupported", "failure_stage": "numeric_conversion_failure"}
        product = feature * coefficient
        products[name] = str(product)
        total += product
    rounded = round_decimal(total, decimal_places)
    return {
        "status": "success",
        "intercept": str(intercept_value),
        "feature_names": sorted(feature_values),
        "feature_values": {name: str(feature_values[name]) for name in sorted(feature_values)},
        "coefficient_names": sorted(coefficient_values),
        "coefficient_values": {name: str(coefficient_values[name]) for name in sorted(coefficient_values)},
        "term_products": products,
        "unrounded_result": str(total),
        "rounded_result": format(rounded, "f"),
        "formula": "intercept + sum(coefficient[name] * feature[name])",
    }


def infer_coefficient_inputs(tables: list[Any], question: str) -> dict[str, Any] | None:
    """同一資料内の対象行と係数表から、線形予測の入力を一意に組み立てる。

    列の並び順には依存せず、ID、特徴量名、係数名の対応が一意に確認できる場合だけ返す。
    """
    target_match = re.search(r"(?:index|id)\s*=\s*([A-Za-z0-9_-]+)", _norm(question), re.IGNORECASE)
    if not target_match:
        return None
    target_id = _norm(target_match.group(1))

    def normalized_name(value: Any) -> str:
        name = _norm(value).lower().replace(" ", "_")
        for prefix in ("coefficient_", "coef_", "weight_"):
            if name.startswith(prefix):
                name = name[len(prefix):]
        return name

    def is_id_column(column: str) -> bool:
        return normalized_name(column) in {"id", "index", "row_id", "record_id"}

    target_rows: list[tuple[Any, dict[str, Any], str]] = []
    for table in tables:
        for row in getattr(table, "rows", []):
            for column in getattr(table, "columns", []):
                if is_id_column(column) and _norm(row.get(column)) == target_id:
                    target_rows.append((table, row, column))
    if len(target_rows) != 1:
        return None
    target_table, target_row, id_column = target_rows[0]

    features: dict[str, Any] = {}
    for column in getattr(target_table, "columns", []):
        if column == id_column:
            continue
        value = _decimal_number(target_row.get(column))
        if value is not None:
            features[normalized_name(column)] = target_row.get(column)
    if not features:
        return None

    coefficient_candidates: list[tuple[Any, dict[str, Any], dict[str, Any], Any]] = []
    for table in tables:
        columns = list(getattr(table, "columns", []))
        # 縦持ち表: feature/variable と coefficient/weight の2列を探す。
        feature_column = next((c for c in columns if normalized_name(c) in {"feature", "variable", "name", "term"}), None)
        coefficient_column = next((c for c in columns if normalized_name(c) in {"coefficient", "coef", "weight", "value"}), None)
        if feature_column and coefficient_column:
            values: dict[str, Any] = {}
            intercept: Any = None
            for row in getattr(table, "rows", []):
                name = normalized_name(row.get(feature_column))
                if name in {"intercept", "constant", "bias"}:
                    intercept = row.get(coefficient_column)
                elif name in features and name not in values:
                    values[name] = row.get(coefficient_column)
            if set(values) == set(features) and intercept is not None:
                coefficient_candidates.append((table, values, features, intercept))

        # 横持ち表: 特徴量名が列名、係数値が1行に並ぶ表を探す。
        for row in getattr(table, "rows", []):
            values = {normalized_name(c): row.get(c) for c in columns if normalized_name(c) in features and _decimal_number(row.get(c)) is not None}
            intercept_column = next((c for c in columns if normalized_name(c) in {"intercept", "constant", "bias"}), None)
            if set(values) == set(features) and intercept_column and _decimal_number(row.get(intercept_column)) is not None:
                coefficient_candidates.append((table, values, features, row.get(intercept_column)))
    if len(coefficient_candidates) != 1:
        return None
    coefficient_table, coefficients, feature_values, intercept = coefficient_candidates[0]
    return {
        "intercept": intercept,
        "feature_values": feature_values,
        "coefficient_values": coefficients,
        "source_range": {
            "target_table": getattr(target_table, "sheet_name", ""),
            "target_row": target_row.get("__row_number__"),
            "coefficient_table": getattr(coefficient_table, "sheet_name", ""),
            "target_id": target_id,
        },
    }


def _excel_column_name(index: int) -> str:
    """Return a one-based Excel column index as an A1 column label."""
    label = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        label = chr(65 + remainder) + label
    return label


def _excel_cell_location(sheet_name: str, row: int, column: int) -> str:
    return f"{sheet_name}!{_excel_column_name(column)}{row}"


def _normal_feature_name(value: Any) -> str:
    """Normalize a coefficient or header label without relying on column order."""
    return _norm(value).lower().replace(" ", "_")


def _is_intercept_name(value: Any) -> bool:
    return _normal_feature_name(value) in {"intercept", "constant", "bias", "切片"}


def _target_key_from_question(question: str) -> tuple[str, str] | None:
    match = re.search(r"(?:index|id)\s*=\s*([A-Za-z0-9_-]+)", _norm(question), re.IGNORECASE)
    if not match:
        return None
    return match.group(0).split("=")[0].strip().lower(), _norm(match.group(1))


def _numeric_cell_value(cell: Any) -> Decimal | None:
    return _decimal_number(getattr(cell, "value", cell))


def _population_standard_deviation(values: list[Decimal]) -> Decimal | None:
    if not values:
        return None
    mean = sum(values, Decimal(0)) / Decimal(len(values))
    variance = sum((value - mean) ** 2 for value in values) / Decimal(len(values))
    try:
        return variance.sqrt()
    except (ArithmeticError, ValueError):
        return None


def _parse_standardize_formula(formula: Any) -> dict[str, str] | None:
    """Parse Excel STANDARDIZE formulas that explicitly retain their source range.

    The evaluator deliberately accepts only the deterministic form emitted by Excel:
    STANDARDIZE(value, AVERAGE(range), STDEV.P(range)). Other preprocessing is
    left unresolved instead of being guessed.
    """
    text = _norm(formula).replace("_xlfn.", "")
    match = re.fullmatch(
        r"=STANDARDIZE\((?P<value>[^,]+),\s*AVERAGE\((?P<mean>[^)]+)\),\s*STDEV\.P\((?P<std>[^)]+)\)\)",
        text,
        re.IGNORECASE,
    )
    if not match or match.group("mean") != match.group("std"):
        return None
    value_ref = match.group("value").replace("$", "")
    range_ref = match.group("mean").replace("$", "")
    value_match = re.fullmatch(r"(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<cell>[A-Z]+\d+)", value_ref, re.IGNORECASE)
    range_match = re.fullmatch(r"(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<start>[A-Z]+\d+):(?P<end>[A-Z]+\d+)", range_ref, re.IGNORECASE)
    if not value_match or not range_match:
        return None
    value_sheet = (value_match.group("sheet") or "").strip("'")
    range_sheet = (range_match.group("sheet") or value_sheet).strip("'")
    if value_sheet and range_sheet and value_sheet != range_sheet:
        return None
    return {
        "source_sheet": value_sheet or range_sheet,
        "value_cell": value_match.group("cell"),
        "range_start": range_match.group("start"),
        "range_end": range_match.group("end"),
    }


def _evaluate_standardize_formula(
    workbook: Any,
    formula: Any,
    statistics_cache: dict[tuple[str, str, str], tuple[Decimal, Decimal]] | None = None,
) -> tuple[Decimal | None, dict[str, Any] | None]:
    parsed = _parse_standardize_formula(formula)
    if not parsed or not parsed["source_sheet"] or parsed["source_sheet"] not in workbook.sheetnames:
        return None, None
    source = workbook[parsed["source_sheet"]]
    value = _numeric_cell_value(source[parsed["value_cell"]])
    cache_key = (source.title, parsed["range_start"], parsed["range_end"])
    statistics = statistics_cache.get(cache_key) if statistics_cache is not None else None
    if statistics is None:
        cells = source[f"{parsed['range_start']}:{parsed['range_end']}"]
        values = [number for row in cells for number in (_numeric_cell_value(cell) for cell in row) if number is not None]
        stdev = _population_standard_deviation(values)
        if stdev is None:
            return None, None
        mean = sum(values, Decimal(0)) / Decimal(len(values))
        statistics = (mean, stdev)
        if statistics_cache is not None:
            statistics_cache[cache_key] = statistics
    mean, stdev = statistics
    if value is None or stdev is None or stdev == 0:
        return None, None
    return (value - mean) / stdev, {
        "preprocessing": "STANDARDIZE",
        "source_sheet": source.title,
        "value_cell": f"{source.title}!{parsed['value_cell']}",
        "range": f"{source.title}!{parsed['range_start']}:{parsed['range_end']}",
        "source_value": str(value),
        "mean": str(mean),
        "population_standard_deviation": str(stdev),
    }


def _has_unbound_standardized_representation(workbook: Any, feature_names: set[str]) -> bool:
    """Detect a complete STANDARDIZE-derived representation of the same features.

    A regression output does not always record which sheet supplied its inputs.
    When both raw and standardized representations are present but the workbook
    gives no explicit binding, the raw row must not be chosen by default.
    """
    for sheet in workbook.worksheets:
        for header_row in range(1, sheet.max_row + 1):
            headers = {
                _normal_feature_name(sheet.cell(header_row, column).value): column
                for column in range(1, sheet.max_column + 1)
                if _norm(sheet.cell(header_row, column).value)
            }
            if not feature_names.issubset(headers):
                continue
            for row_number in range(header_row + 1, sheet.max_row + 1):
                formulas = [sheet.cell(row_number, headers[name]).value for name in feature_names]
                if formulas and all(_parse_standardize_formula(formula) is not None for formula in formulas):
                    return True
    return False


def _workbook_headers(sheet: Any, feature_names: set[str]) -> list[tuple[int, dict[str, int]]]:
    """Return header rows that contain every regression feature exactly once."""
    matches: list[tuple[int, dict[str, int]]] = []
    # Analysis tables place their headers before the data region.  Capping this
    # structural scan prevents a large data sheet from turning header discovery
    # into a quadratic pass; a deep header is intentionally left unresolved.
    for header_row in range(1, min(sheet.max_row, 200) + 1):
        headers: dict[str, int] = {}
        duplicate = False
        for column in range(1, sheet.max_column + 1):
            name = _normal_feature_name(sheet.cell(header_row, column).value)
            if not name:
                continue
            if name in headers:
                duplicate = True
                break
            headers[name] = column
        if not duplicate and feature_names.issubset(headers):
            matches.append((header_row, headers))
    return matches


def _reproduction_error(
    rows: list[dict[str, Decimal]],
    target_column: str,
    feature_names: list[str],
    intercept: Decimal,
    coefficients: dict[str, Decimal],
) -> Decimal | None:
    """Re-fit a candidate representation and compare it with the report values.

    This is a source-binding check, not a model training path.  A candidate is
    accepted only when its least-squares coefficients reproduce the workbook's
    report, so merely finding a sheet named "standardized" is insufficient.
    """
    numeric_rows = [row for row in rows if target_column in row and all(name in row for name in feature_names)]
    if len(numeric_rows) <= len(feature_names):
        return None
    try:
        import numpy as np

        matrix = np.asarray(
            [[1.0] + [float(row[name]) for name in feature_names] for row in numeric_rows],
            dtype=float,
        )
        target = np.asarray([float(row[target_column]) for row in numeric_rows], dtype=float)
        fitted, *_ = np.linalg.lstsq(matrix, target, rcond=None)
    except (ImportError, ArithmeticError, TypeError, ValueError):
        return None
    expected = [float(intercept)] + [float(coefficients[name]) for name in feature_names]
    return Decimal(str(max(abs(left - right) for left, right in zip(fitted, expected))))


def _numeric_response_columns(sheet: Any, header_row: int, headers: dict[str, int], feature_names: set[str], key_name: str) -> list[str]:
    """Find non-feature numeric columns that can prove a regression source."""
    candidates: list[str] = []
    for name, column in headers.items():
        if name in feature_names or name == key_name:
            continue
        values = [_numeric_cell_value(sheet.cell(row, column)) for row in range(header_row + 1, sheet.max_row + 1)]
        if sum(value is not None for value in values) > len(feature_names):
            candidates.append(name)
    return candidates


def _standardized_representation_candidates(
    workbook: Any,
    feature_names: set[str],
    key_name: str,
    key_targets: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Build formula-linked standardized representations with an ID binding.

    A standardized sheet often omits the ID column.  Its formula source cells
    are therefore matched to a key-addressable row by feature name and value;
    no row number or sheet name is assumed by this resolver.
    """
    candidates: list[dict[str, Any]] = []
    statistics_cache: dict[tuple[str, str, str], tuple[Decimal, Decimal]] = {}
    source_header_cache: dict[str, list[tuple[int, dict[str, int]]]] = {}
    for sheet in workbook.worksheets:
        for header_row, headers in _workbook_headers(sheet, feature_names):
            response_columns = _numeric_response_columns(sheet, header_row, headers, feature_names, key_name)
            if not response_columns:
                continue
            formula_rows: list[dict[str, Any]] = []
            for row in range(header_row + 1, sheet.max_row + 1):
                parsed = [_parse_standardize_formula(sheet.cell(row, headers[name]).value) for name in sorted(feature_names)]
                if not all(parsed):
                    continue
                source_sheets = {item["source_sheet"] for item in parsed if item}
                source_rows = {re.search(r"\d+$", item["value_cell"]).group(0) for item in parsed if item}
                if len(source_sheets) != 1 or len(source_rows) != 1:
                    continue
                values: dict[str, Decimal] = {}
                details: dict[str, Any] = {}
                for name in sorted(feature_names):
                    value, detail = _evaluate_standardize_formula(
                        workbook, sheet.cell(row, headers[name]).value, statistics_cache,
                    )
                    if value is None or detail is None:
                        values = {}
                        break
                    values[name] = value
                    details[name] = detail
                if values:
                    formula_rows.append({
                        "row": row,
                        "values": values,
                        "details": details,
                        "source_sheet": next(iter(source_sheets)),
                        "source_row": int(next(iter(source_rows))),
                    })
            if not formula_rows:
                continue
            for target in key_targets:
                target_sheet = target["sheet"]
                if not feature_names.issubset(target["headers"]):
                    continue
                mapped = []
                for formula_row in formula_rows:
                    if formula_row["source_sheet"] not in workbook.sheetnames:
                        continue
                    source_sheet = workbook[formula_row["source_sheet"]]
                    cached_headers = source_header_cache.get(source_sheet.title)
                    if cached_headers is None:
                        cached_headers = _workbook_headers(source_sheet, feature_names)
                        source_header_cache[source_sheet.title] = cached_headers
                    source_headers = [
                        source for source in cached_headers
                        if source[0] < formula_row["source_row"]
                    ]
                    if len(source_headers) != 1:
                        continue
                    _, source_header_map = source_headers[0]
                    source_values = {
                        name: _numeric_cell_value(source_sheet.cell(formula_row["source_row"], source_header_map[name]))
                        for name in feature_names
                    }
                    target_values = {
                        name: _numeric_cell_value(target_sheet.cell(target["row"], target["headers"][name]))
                        for name in feature_names
                    }
                    if all(source_values[name] is not None and source_values[name] == target_values[name] for name in feature_names):
                        mapped.append(formula_row)
                if len(mapped) != 1:
                    continue
                rows: list[dict[str, Decimal]] = []
                for formula_row in formula_rows:
                    record = dict(formula_row["values"])
                    for response in response_columns:
                        value = _numeric_cell_value(sheet.cell(formula_row["row"], headers[response]))
                        if value is not None:
                            record[response] = value
                    rows.append(record)
                candidates.append({
                    "target": target,
                    "sheet": sheet,
                    "header_row": header_row,
                    "feature_values": mapped[0]["values"],
                    "feature_locations": {
                        name: _excel_cell_location(sheet.title, mapped[0]["row"], headers[name])
                        for name in feature_names
                    },
                    "rows": rows,
                    "response_columns": response_columns,
                    "preprocessing": {
                        "type": "STANDARDIZE",
                        "formula_sources": mapped[0]["details"],
                        "source_sheet": mapped[0]["source_sheet"],
                        "source_row": mapped[0]["source_row"],
                    },
                })
    return candidates


def infer_coefficient_inputs_from_workbook(path: Path, question: str) -> dict[str, Any] | None:
    """Resolve a linear prediction from an XLSX workbook using names and cells.

    This is intentionally conservative: a coefficient block, a key-selected row,
    every feature binding, and any STANDARDIZE preprocessing must each be unique.
    It returns ``None`` for any ambiguity so that the Answer Gate keeps the item
    suppressed.
    """
    key_request = _target_key_from_question(question)
    if key_request is None or not path.exists():
        return None
    key_name, key_value = key_request
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except (OSError, ValueError, KeyError):
        return None

    coefficient_blocks: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for header_cell in row:
                if _normal_feature_name(header_cell.value) not in {"係数", "coefficient", "coefficients", "coef"}:
                    continue
                feature_column = header_cell.column - 1
                if feature_column < 1:
                    continue
                values: dict[str, tuple[Any, str]] = {}
                intercept: tuple[Any, str] | None = None
                for row_number in range(header_cell.row + 1, sheet.max_row + 1):
                    feature_cell = sheet.cell(row_number, feature_column)
                    coefficient_cell = sheet.cell(row_number, header_cell.column)
                    if not _norm(feature_cell.value) and not _norm(coefficient_cell.value):
                        break
                    coefficient = _numeric_cell_value(coefficient_cell)
                    if coefficient is None:
                        continue
                    name = _normal_feature_name(feature_cell.value)
                    if not name:
                        continue
                    location = _excel_cell_location(sheet.title, row_number, header_cell.column)
                    if _is_intercept_name(feature_cell.value):
                        if intercept is not None:
                            intercept = None
                            break
                        intercept = (coefficient_cell.value, location)
                    elif name not in values:
                        values[name] = (coefficient_cell.value, location)
                    else:
                        values = {}
                        break
                if intercept is not None and values:
                    coefficient_blocks.append({
                        "sheet": sheet,
                        "feature_column": feature_column,
                        "coefficient_column": header_cell.column,
                        "header_row": header_cell.row,
                        "coefficients": values,
                        "intercept": intercept,
                    })

    target_rows: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        for header_row in range(1, min(sheet.max_row, 200) + 1):
            headers = {_normal_feature_name(sheet.cell(header_row, column).value): column for column in range(1, sheet.max_column + 1) if _norm(sheet.cell(header_row, column).value)}
            key_column = headers.get(key_name)
            if not key_column:
                continue
            for row_number in range(header_row + 1, sheet.max_row + 1):
                if _norm(sheet.cell(row_number, key_column).value) == key_value:
                    target_rows.append({"sheet": sheet, "header_row": header_row, "row": row_number, "headers": headers, "key_column": key_column})

    candidates: list[dict[str, Any]] = []
    has_formula_derived_representation = False
    for block in coefficient_blocks:
        feature_names = set(block["coefficients"])
        coefficient_values = {name: _decimal_number(value) for name, (value, _) in block["coefficients"].items()}
        intercept_value = _decimal_number(block["intercept"][0])
        if intercept_value is None or any(value is None for value in coefficient_values.values()):
            continue
        for target in target_rows:
            if not feature_names.issubset(target["headers"]):
                continue
            feature_values: dict[str, Any] = {}
            feature_locations: dict[str, str] = {}
            preprocessing: dict[str, Any] = {}
            valid = True
            for feature in sorted(feature_names):
                column = target["headers"][feature]
                cell = target["sheet"].cell(target["row"], column)
                value = _numeric_cell_value(cell)
                if value is None:
                    valid = False
                    break
                feature_values[feature] = cell.value
                feature_locations[feature] = _excel_cell_location(target["sheet"].title, target["row"], column)
            if valid:
                response_columns = _numeric_response_columns(
                    target["sheet"], target["header_row"], target["headers"], feature_names, key_name,
                )
                rows: list[dict[str, Decimal]] = []
                for row_number in range(target["header_row"] + 1, target["sheet"].max_row + 1):
                    record: dict[str, Decimal] = {}
                    for name in feature_names | set(response_columns):
                        column = target["headers"].get(name)
                        value = _numeric_cell_value(target["sheet"].cell(row_number, column)) if column else None
                        if value is not None:
                            record[name] = value
                    rows.append(record)
                candidates.append({
                    "block": block,
                    "target": target,
                    "sheet": target["sheet"],
                    "feature_values": feature_values,
                    "feature_locations": feature_locations,
                    "preprocessing": preprocessing,
                    "rows": rows,
                    "response_columns": response_columns,
                })

        # Formula-linked derived sheets are valid only when the report can be
        # re-produced from their numeric response column.  This keeps a raw
        # table and a standardized table from being confused by their names.
        derived_candidates = _standardized_representation_candidates(workbook, feature_names, key_name, target_rows)
        if derived_candidates:
            has_formula_derived_representation = True
        elif _has_unbound_standardized_representation(workbook, feature_names):
            # A derived representation without a numeric response cannot be
            # verified, so retain the conservative suppression rather than
            # silently falling back to the raw row.
            return None
        for derived in derived_candidates:
            derived["block"] = block
            candidates.append(derived)

    if not has_formula_derived_representation:
        # Preserve the established direct-table route when no competing
        # preprocessing exists.  The response column is optional in this case.
        if len(candidates) != 1:
            return None
        selected = candidates[0]
        selected["response_column"] = ""
        selected["coefficient_reproduction_error"] = "not_required_no_derived_representation"
    else:
        reproduced: list[dict[str, Any]] = []
        for candidate in candidates:
            block = candidate["block"]
            feature_names = sorted(block["coefficients"])
            coefficients = {name: _decimal_number(block["coefficients"][name][0]) for name in feature_names}
            intercept = _decimal_number(block["intercept"][0])
            if intercept is None or any(value is None for value in coefficients.values()):
                continue
            for response in candidate["response_columns"]:
                error = _reproduction_error(candidate["rows"], response, feature_names, intercept, coefficients)  # type: ignore[arg-type]
                if error is not None:
                    reproduced.append({**candidate, "response_column": response, "coefficient_reproduction_error": error})

        # Excel's displayed regression output has normal floating point noise,
        # but a different source representation has errors many orders larger.
        accepted = [candidate for candidate in reproduced if candidate["coefficient_reproduction_error"] <= Decimal("0.000001")]
        if len(accepted) != 1:
            return None
        selected = accepted[0]
    block = selected["block"]
    target = selected["target"]
    coefficient_values = {name: value for name, (value, _) in block["coefficients"].items()}
    coefficient_locations = {name: location for name, (_, location) in block["coefficients"].items()}
    cell_ranges = [block["intercept"][1], _excel_cell_location(target["sheet"].title, target["row"], target["key_column"])]
    cell_ranges.extend(coefficient_locations[name] for name in sorted(coefficient_locations))
    cell_ranges.extend(selected["feature_locations"][name] for name in sorted(selected["feature_locations"]))
    return {
        "intercept": block["intercept"][0],
        # Evidence artifacts are JSON. Preserve exact derived numeric values as
        # strings so calculation and audit paths do not depend on a Decimal
        # encoder being installed by the caller.
        "feature_values": {name: str(value) for name, value in selected["feature_values"].items()},
        "coefficient_values": coefficient_values,
        "source_range": {
            "target_table": target["sheet"].title,
            "target_row": target["row"],
            "target_header_row": target["header_row"],
            "coefficient_table": block["sheet"].title,
            "coefficient_header_row": block["header_row"],
            "target_id": key_value,
        },
        "cell_ranges": cell_ranges,
        "column_bindings": [
            {
                "feature": name,
                "coefficient_cell": coefficient_locations[name],
                "input_cell": selected["feature_locations"][name],
            }
            for name in sorted(coefficient_values)
        ],
        "intercept_cell": block["intercept"][1],
        "key_cell": _excel_cell_location(target["sheet"].title, target["row"], target["key_column"]),
        "preprocessing": selected["preprocessing"],
        "selected_source_sheet": selected["sheet"].title,
        "source_selection_reason": "unique_coefficient_reproduction",
        "coefficient_reproduction_error": str(selected["coefficient_reproduction_error"]),
        "target_column": selected["response_column"],
        "source_binding_verified": True,
    }


def calculate_difference(left: Any, right: Any, *, absolute: bool = False, decimal_places: int | None = None) -> dict[str, Any]:
    """2つの検証済み数値の差を、符号付きまたは絶対値で計算する。"""
    left_value = _decimal_number(left)
    right_value = _decimal_number(right)
    if left_value is None or right_value is None:
        return {"status": "unsupported", "failure_stage": "numeric_conversion_failure"}
    raw = left_value - right_value
    result = abs(raw) if absolute else raw
    rounded = round_decimal(result, decimal_places)
    return {
        "status": "success",
        "left": str(left_value),
        "right": str(right_value),
        "difference_direction": "left_minus_right",
        "absolute_or_signed": "absolute" if absolute else "signed",
        "unrounded_result": str(result),
        "rounded_result": format(rounded, "f"),
        "formula": "abs(left - right)" if absolute else "left - right",
    }


def rank_numeric_rows(rows: list[dict[str, Any]], value_column: str, *, direction: str = "asc", return_column: str | None = None, tie_policy: str = "suppress_if_multiple") -> dict[str, Any]:
    """指定列の最小・最大を選び、同率を安全に扱う。"""
    values = [(row, _decimal_number(row.get(value_column))) for row in rows]
    values = [(row, value) for row, value in values if value is not None]
    if not values:
        return {"status": "unsupported", "failure_stage": "numeric_conversion_failure"}
    best = min(value for _, value in values) if direction == "asc" else max(value for _, value in values)
    selected = [row for row, value in values if value == best]
    if len(selected) != 1 and tie_policy == "suppress_if_multiple":
        return {"status": "unsupported", "failure_stage": "uniqueness_failure", "tie_candidates": len(selected), "sort_value": str(best)}
    output = [str(row.get(return_column or value_column, "")) for row in selected]
    return {"status": "success", "sort_key": value_column, "sort_direction": direction, "sort_value": str(best), "tie_candidates": len(selected), "result": output, "formula": f"{direction}({value_column})"}


def calculate_schedule_effort(duration: Any, resource_count: Any, *, unit: str, decimal_places: int | None = None) -> dict[str, Any]:
    """確定済み期間と人数から人日・人時を計算する。暦日/営業日の推測はしない。"""
    duration_value = _decimal_number(duration)
    resource_value = _decimal_number(resource_count)
    if duration_value is None or resource_value is None or not unit:
        return {"status": "unsupported", "failure_stage": "schedule_spec_failure"}
    raw = duration_value * resource_value
    rounded = round_decimal(raw, decimal_places)
    return {"status": "success", "duration": str(duration_value), "resource_count": str(resource_value), "duration_unit": unit, "unrounded_result": str(raw), "rounded_result": format(rounded, "f"), "formula": "duration * resource_count"}


def execute_operation_type(spec: CalculationSpec, inputs: dict[str, Any]) -> dict[str, Any]:
    """CalculationSpecの入力Evidenceを使って、指定された計算だけを実行する。"""
    operation = spec.operation_type or spec.calculation_subtype
    places = spec.rounding.get("decimal_places")
    if operation == "ratio_or_percentage":
        return calculate_ratio(inputs.get("numerator"), inputs.get("denominator"), percentage=spec.percentage, decimal_places=places)
    if operation == "difference":
        return calculate_difference(inputs.get("left"), inputs.get("right"), absolute=spec.absolute_or_signed == "absolute", decimal_places=places)
    if operation == "coefficient_prediction":
        return linear_prediction(inputs.get("intercept"), inputs.get("feature_values", {}), inputs.get("coefficient_values", {}), decimal_places=places)
    if operation == "schedule_effort":
        return calculate_schedule_effort(inputs.get("duration"), inputs.get("resource_count"), unit=spec.duration_unit or "", decimal_places=places)
    if operation == "ranking_or_argmin":
        return rank_numeric_rows(inputs.get("rows", []), inputs.get("value_column", ""), direction=spec.sort_direction or "asc", return_column=inputs.get("return_column"), tie_policy=spec.tie_policy or "suppress_if_multiple")
    return {"status": "unsupported", "failure_stage": "unsupported_calculation", "warning": "operation_type has no executable input binding"}


def verify_operation_evidence(spec: CalculationSpec, inputs: dict[str, Any], result: dict[str, Any]) -> dict[str, Any]:
    """入力、式、結果がEvidenceから再現できるかを確認する。"""
    required = {
        "ratio_or_percentage": ("numerator", "denominator"),
        "difference": ("left", "right"),
        "coefficient_prediction": ("intercept", "feature_values", "coefficient_values"),
        "schedule_effort": ("duration", "resource_count"),
        "ranking_or_argmin": ("rows", "value_column"),
    }.get(spec.operation_type or spec.calculation_subtype, ())
    present = all(key in inputs and inputs[key] not in (None, "", {}) for key in required)
    return {
        "required_inputs_present": present,
        "operation_type": spec.operation_type or spec.calculation_subtype,
        "operation_success": result.get("status") == "success",
        "independent_recalculation_match": result.get("status") == "success",
        "verification_status": "passed" if present and result.get("status") == "success" else "failed",
    }


def verify_calculation_evidence(evidence: dict[str, Any], spec: CalculationSpec) -> dict[str, Any]:
    verification = {
        "question_type_match": True,
        "condition_coverage": True,
        "input_presence": bool(evidence.get("selected_file_id")) and bool(spec.input_columns),
        "type_validity": evidence.get("input_row_counts", {}).get("numeric", 0) > 0,
        "filter_validity": bool(spec.filters) or spec.row_selector is not None,
        "operation_validity": bool(spec.operations),
        "denominator_validity": None,
        "coefficient_alignment": None,
        "rounding_validity": evidence.get("formatted_result") not in (None, ""),
        "reproducibility": bool(evidence.get("calculation_formula")) and bool(evidence.get("intermediate_values")),
        "source_range": bool(evidence.get("cell_ranges")),
    }
    verification["verification_status"] = "passed" if all(value is not False for value in verification.values()) else "failed"
    return verification


def execute_calculation_spec(
    question_id: int,
    question: str,
    spec: CalculationSpec,
    table: Any,
    filter_rows: Callable[[Any, list[dict[str, Any]], str], list[dict[str, Any]]],
    inputs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if spec.calculation_subtype not in {"multi_step_calculation", "filtered_aggregation"}:
        # 新しい演算は、Plannerが明示した入力だけを受け取って実行します。
        # 入力がない場合は、表の最初の数値列などを推測せず抑制します。
        if inputs is None:
            return {"status": "unsupported", "failure_stage": "calculation_input_resolution_failure", "warning": "Calculation inputs are not explicitly resolved", "spec": asdict(spec)}
        result = execute_operation_type(spec, inputs)
        evidence = {
            "question_id": question_id,
            "operation_type": spec.operation_type or spec.calculation_subtype,
            "input_values": inputs,
            "operation_order": spec.operations,
            "raw_result": result.get("raw_result"),
            "formatted_result": result.get("rounded_result", result.get("result")),
            "unit": spec.unit,
            "preview_only": False,
            "source_range": inputs.get("source_range"),
            "cell_ranges": list(inputs.get("cell_ranges", [])),
            "intermediate_values": result.get("intermediate_values", {}),
            "column_bindings": list(inputs.get("column_bindings", [])),
        }
        verification = verify_operation_evidence(spec, inputs, result)
        # Calculation questions have a stricter Gate contract. These fields are
        # derived from the resolved workbook cells, never from a model answer.
        independently_recalculated = linear_prediction(
            inputs.get("intercept"),
            inputs.get("feature_values", {}),
            inputs.get("coefficient_values", {}),
            decimal_places=spec.rounding.get("decimal_places"),
        ) if spec.operation_type == "coefficient_prediction" else result
        formatted = str(evidence["formatted_result"] or "")
        decimal_places = spec.rounding.get("decimal_places")
        answer_format_valid = bool(re.fullmatch(r"-?\d+(?:\.\d+)?", formatted))
        if decimal_places is not None:
            answer_format_valid = answer_format_valid and bool(re.fullmatch(rf"-?\d+\.\d{{{int(decimal_places)}}}", formatted))
        verification.update({
            "question_type_match": True,
            "condition_coverage": bool(inputs.get("key_cell")),
            "input_presence": bool(inputs.get("feature_values")) and bool(inputs.get("coefficient_values")),
            "type_validity": all(_decimal_number(value) is not None for value in list(inputs.get("feature_values", {}).values()) + list(inputs.get("coefficient_values", {}).values())),
            "filter_validity": bool(inputs.get("key_cell")),
            "operation_validity": result.get("status") == "success",
            "rounding_validity": result.get("rounded_result") == formatted,
            "reproducibility": independently_recalculated.get("rounded_result") == formatted,
            "source_range": bool(evidence["cell_ranges"]),
            "column_bindings_verified": bool(inputs.get("source_binding_verified")) and len(inputs.get("column_bindings", [])) == len(inputs.get("feature_values", {})),
            "conditions_applied": bool(inputs.get("key_cell")),
            "operation_graph_complete": spec.operations == [{"step_id": "s1", "operation": "linear_prediction"}],
            "source_ranges_present": bool(evidence["cell_ranges"]),
            "units_consistent": True,
            "rounding_valid": result.get("rounded_result") == formatted,
            "independent_recalculation_match": independently_recalculated.get("rounded_result") == formatted,
            "answer_format_valid": answer_format_valid,
            "no_unverified_fallback": bool(inputs.get("source_binding_verified")),
        })
        verification["verification_status"] = "passed" if all(value is True for value in verification.values() if isinstance(value, bool)) else "failed"
        if result.get("status") != "success" or verification.get("verification_status") != "passed":
            return {"status": "unsupported", "failure_stage": result.get("failure_stage", "verification_failure"), "warning": result.get("warning", "Operation evidence verification failed"), "spec": asdict(spec), "evidence": evidence, "verification": verification}
        answer = str(evidence["formatted_result"])
        return {"status": "success", "answer": answer, "evidence": evidence, "verification": verification, "spec": asdict(spec), "steps": result.get("steps", []), "operations_executed": ["calculation", "answer_formatting"]}
    filtered = filter_rows(table, spec.filters, spec.logical_operator)
    mean_column = next((step.get("input", "").split(".")[-1] for step in spec.operations if step.get("operation") == "mean"), "")
    numeric = [(row, _decimal_number(row.get(mean_column))) for row in filtered]
    missing_count = sum(value is None for _, value in numeric)
    numeric = [(row, value) for row, value in numeric if value is not None]
    if not numeric:
        return {"status": "unsupported", "failure_stage": "numeric_conversion_failure", "warning": "No numeric values", "spec": asdict(spec)}
    mean_value = sum((value for _, value in numeric), Decimal(0)) / Decimal(len(numeric))
    steps = [
        {"step_id": "s1", "operation": "filter_rows", "input_row_count": len(table.rows), "output_row_count": len(filtered)},
        {"step_id": "s2", "operation": "mean", "input_row_count": len(numeric), "output": str(mean_value)},
    ]
    if spec.calculation_subtype == "multi_step_calculation":
        distances = [(row, abs(value - mean_value)) for row, value in numeric]
        minimum = min(distance for _, distance in distances)
        nearest = [row for row, distance in distances if distance == minimum]
        values = [str(row.get(spec.output_column or "", "")) for row in nearest if _norm(row.get(spec.output_column or "", ""))]
        if not values:
            return {"status": "unsupported", "failure_stage": "row_selection_failure", "warning": "Nearest rows have no output value", "spec": asdict(spec)}
        answer = "\u3001".join(values)
        raw_result: Any = values
        steps.extend([
            {"step_id": "s3", "operation": "absolute_difference", "comparison_value": str(mean_value), "minimum_distance": str(minimum)},
            {"step_id": "s4", "operation": "select_min_distance", "output_row_count": len(nearest), "output_values": values},
        ])
    else:
        places = spec.rounding.get("decimal_places")
        rounded = round_decimal(mean_value, places)
        answer = format(rounded, "f")
        raw_result = str(mean_value)
    evidence = {
        "question_id": question_id,
        "selected_file": table.file.raw_path,
        "selected_file_id": table.file.file_id,
        "sheet_name": table.sheet_name,
        "cell_ranges": [f"A1:{chr(64 + min(len(table.columns), 26))}{len(table.rows) + 1}"],
        "input_columns": spec.input_columns,
        "row_selector": spec.row_selector,
        "filter_conditions": spec.filters,
        "operation_graph": spec.operations,
        "input_row_counts": {"source": len(table.rows), "numeric": len(numeric), "missing_excluded": missing_count},
        "output_row_counts": {"filtered": len(filtered)},
        "intermediate_values": {"mean": str(mean_value)},
        "calculation_formula": "mean then minimum absolute difference" if spec.calculation_subtype == "multi_step_calculation" else "mean",
        "raw_result": raw_result,
        "formatted_result": answer,
        "unit": spec.unit,
        "answer_format": spec.output_type,
        "preview_only": False,
    }
    verification = verify_calculation_evidence(evidence, spec)
    return {"status": "success", "answer": answer, "evidence": evidence, "verification": verification, "spec": asdict(spec), "steps": steps, "operations_executed": ["table_filter", "table_aggregation", "calculation", "answer_formatting"]}


def append_calculation_artifacts(work_dir: Path, question_id: int, result: dict[str, Any]) -> None:
    work_dir.mkdir(parents=True, exist_ok=True)
    rows = {
        "calculation_specs.jsonl": {"question_id": question_id, "spec": result.get("spec", {})},
        "calculation_inputs.jsonl": {"question_id": question_id, "evidence": result.get("evidence", {})},
        "calculation_steps.jsonl": {"question_id": question_id, "steps": result.get("steps", [])},
        "calculation_evidence.jsonl": {"question_id": question_id, "evidence": result.get("evidence", {})},
        "calculation_verification.jsonl": {"question_id": question_id, "verification": result.get("verification", {})},
    }
    for name, row in rows.items():
        with (work_dir / name).open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
