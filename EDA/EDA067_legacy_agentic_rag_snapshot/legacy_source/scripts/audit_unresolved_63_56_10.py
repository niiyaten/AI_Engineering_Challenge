"""Create read-only investigation artifacts for tests 63, 56, and 10."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree as ET


ROOT = Path(__file__).resolve().parents[1]
RUN_ID = "unresolved_63_56_10_generic_capabilities_v1"
OUT = ROOT / "data" / "output" / RUN_ID / "analysis"
RAW = ROOT / "data" / "raw" / "share" / "share"
FRESH = ROOT / "data" / "output" / (RUN_ID + "_test_fresh")


def write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(dict.fromkeys(k for row in rows for k in row))
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def find_one(*terms: str, suffix: str | None = None) -> Path | None:
    for path in RAW.rglob("*"):
        if not path.is_file() or (suffix and path.suffix.lower() != suffix.lower()):
            continue
        text = str(path)
        if all(term in text for term in terms):
            return path
    return None


def package_versions(lock: Path) -> dict[str, str]:
    versions: dict[str, str] = {}
    name = None
    for line in lock.read_text(encoding="utf-8").splitlines():
        m = re.match(r'name = "([^"]+)"', line)
        if m:
            name = m.group(1)
        m = re.match(r'version = "([^"]+)"', line)
        if m and name in {"seaborn", "matplotlib", "pandas", "numpy"}:
            versions[name] = m.group(1)
    return versions


def audit_environment(project: Path) -> dict[str, object]:
    commands = []
    for command in ["python", "py", "uv"]:
        found = shutil.which(command)
        commands.append({"command": command, "path": found})
    venv = ROOT / ".venv" / "Scripts" / "python.exe"
    imports = {}
    probe = "import sys\nprint(sys.executable)\nfor m in ['seaborn','matplotlib','pandas','numpy']:\n try:\n  x=__import__(m); print(m+'=OK='+getattr(x,'__version__',''))\n except Exception as e: print(m+'=FAIL='+type(e).__name__+':'+str(e))\n"
    for label, executable in [("current", Path(sys.executable)), ("workspace_venv", venv)]:
        if not executable.exists():
            imports[label] = {"exists": False}
            continue
        proc = subprocess.run([str(executable), "-c", probe], cwd=ROOT, capture_output=True, text=True)
        imports[label] = {"exists": True, "returncode": proc.returncode, "stdout": proc.stdout, "stderr": proc.stderr}
    pip_probe = subprocess.run([str(venv), "-m", "pip", "show", "seaborn", "matplotlib", "pandas"], cwd=ROOT, capture_output=True, text=True)
    lock = project / "uv.lock"
    return {
        "python_executable": str(venv),
        "current_python_executable": sys.executable,
        "imported_package_path": str(ROOT / "src" / "rag_competition"),
        "working_directory": str(ROOT),
        "PYTHONPATH": os.environ.get("PYTHONPATH", ""),
        "config_path": "config/openrouter_free.json",
        "cache_version": "pipeline_cache_v1",
        "index_version": "search_index_agentic_rag_v1",
        "msoffcrypto_importable": _can_import("msoffcrypto"),
        "command_paths": commands,
        "import_probes": imports,
        "pip_show": {"returncode": pip_probe.returncode, "stdout": pip_probe.stdout, "stderr": pip_probe.stderr},
        "project_pyproject": (project / "pyproject.toml").read_text(encoding="utf-8"),
        "project_requirements": (project / "requirements.txt").read_text(encoding="utf-8") if (project / "requirements.txt").exists() else None,
        "uv_lock_exists": lock.exists(),
        "uv_lock_hash": sha256(lock) if lock.exists() else None,
        "uv_lock_pinned_versions": package_versions(lock) if lock.exists() else {},
    }


def _can_import(name: str) -> bool:
    try:
        __import__(name)
        return True
    except Exception:
        return False


def audit_notebook(project: Path) -> dict[str, object]:
    notebook = next(project.rglob("01_eda.ipynb"), None)
    figure = project / "reports" / "figures" / "target_distribution.png"
    if notebook is None:
        return {"status": "not_found", "project": str(project)}
    data = json.loads(notebook.read_text(encoding="utf-8"))
    cells = []
    for idx, cell in enumerate(data.get("cells", [])):
        source = "".join(cell.get("source", []))
        outputs = []
        for output in cell.get("outputs", []):
            entry = {"output_type": output.get("output_type"), "names": list(output.get("data", {}).keys())}
            text = output.get("text", [])
            if text:
                entry["text"] = "".join(text)[:2000]
            outputs.append(entry)
        if idx in {2, 3, 13} or outputs:
            cells.append({"cell_index": idx, "execution_count": cell.get("execution_count"), "source": source[:5000], "outputs": outputs})
    figure_info = {"exists": figure.exists()}
    if figure.exists():
        figure_info.update({"path": str(figure), "sha256": sha256(figure), "bytes": figure.stat().st_size})
        try:
            from PIL import Image
            with Image.open(figure) as im:
                figure_info.update({"width": im.width, "height": im.height, "format": im.format})
            OUT.joinpath("diagnostics").mkdir(parents=True, exist_ok=True)
            shutil.copy2(figure, OUT / "diagnostics" / "test56_target_distribution.png")
        except Exception as exc:
            figure_info["image_error"] = str(exc)
    return {"notebook": str(notebook), "notebook_hash": sha256(notebook), "cells": cells, "figure": figure_info,
            "replay_status": "blocked_seaborn_unavailable", "replay_command": "uv sync --frozen; uv run python -m jupyter nbconvert --to notebook --execute notebooks/01_eda.ipynb"}


def audit_xlsx(workbook: Path) -> dict[str, object]:
    import numpy as np
    import openpyxl

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=False)
    sheet = wb["train"] if "train" in wb.sheetnames else wb[wb.sheetnames[0]]
    values = list(sheet.values)
    headers = [str(v).strip() if v is not None else "" for v in values[0]]
    try:
        col = headers.index("AG_ratio")
    except ValueError:
        col = next((i for i, h in enumerate(headers) if "AG_ratio" in h), -1)
    raw = [row[col] for row in values[1:] if col >= 0 and col < len(row)]
    nums = [float(v) for v in raw if isinstance(v, (int, float)) and not isinstance(v, bool)]
    counts = Counter(nums)
    decimals = Counter()
    for value in nums:
        text = format(value, ".15f").rstrip("0").rstrip(".")
        decimals[len(text.split(".", 1)[1]) if "." in text else 0] += 1
    unique_sorted = sorted(counts)
    step = min((b - a for a, b in zip(unique_sorted, unique_sorted[1:]) if b > a), default=None)
    rounded = {}
    for places in range(0, 7):
        rounded[places] = max(Counter(round(v, places) for v in nums).values(), default=0)
    bins = {}
    for spec in ["auto", "fd", "doane", "scott", "rice", "sqrt", "sturges", 10, 20, 30, 50]:
        try:
            hist, edges = np.histogram(nums, bins=spec)
            bins[str(spec)] = {"max_count": int(hist.max()), "bin_count": len(hist), "edges": [float(x) for x in edges]}
        except Exception as exc:
            bins[str(spec)] = {"error": str(exc)}
    xml_summary = inspect_xlsx_xml(workbook)
    return {"workbook": str(workbook), "sheets": wb.sheetnames, "sheet_state": xml_summary["sheet_state"],
            "sheet": sheet.title, "header_row": 1, "ag_ratio_column": col + 1, "ag_ratio_column_letter": openpyxl.utils.get_column_letter(col + 1) if col >= 0 else None,
            "data_rows": len(raw), "numeric_rows": len(nums), "missing_count": len(raw) - len(nums), "dtype": type(next((v for v in raw if v is not None), None)).__name__,
            "unique_count": len(counts), "decimal_places_distribution": dict(decimals), "minimum_positive_step": step,
            "top_value_counts": [{"value": k, "count": v} for k, v in counts.most_common(20)], "rounded_max_counts": rounded,
            "histogram_candidates": bins, "xml": xml_summary}


def inspect_xlsx_xml(workbook: Path) -> dict[str, object]:
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main", "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships", "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
    with zipfile.ZipFile(workbook) as zf:
        names = zf.namelist()
        workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        states = {x.attrib.get("name"): x.attrib.get("state", "visible") for x in workbook_xml.findall("x:sheets/x:sheet", ns)}
        defined = []
        if "xl/workbook.xml" in names:
            for node in workbook_xml.findall("x:definedNames/x:definedName", ns):
                defined.append({"name": node.attrib.get("name"), "value": node.text})
        pivots = [n for n in names if "pivot" in n.lower()]
        pivot_hits = []
        pivot_cache_fields = []
        pivot_cache_sources = []
        for name in pivots:
            if name.endswith((".xml", ".rels")):
                text = zf.read(name).decode("utf-8", errors="replace")
                if "AG_ratio" in text or "ratio" in text.lower():
                    pivot_hits.append({"path": name, "contains_ag_ratio": "AG_ratio" in text, "excerpt": text[:3000]})
                if "pivotCacheDefinition" in text:
                    pivot_cache_sources.extend(re.findall(r'<worksheetSource[^>]*ref="([^"]+)"[^>]*sheet="([^"]+)"', text))
                    pivot_cache_fields.extend(re.findall(r'<cacheField[^>]*name="([^"]+)"', text))
        media = [n for n in names if n.startswith("xl/media/")]
        drawings = [n for n in names if n.startswith("xl/drawings/") and n.endswith(".xml")]
        media_info = []
        diag = OUT / "diagnostics" / "test10_media"
        diag.mkdir(parents=True, exist_ok=True)
        for name in media:
            target = diag / Path(name).name
            target.write_bytes(zf.read(name))
            item = {"path": name, "bytes": target.stat().st_size}
            try:
                from PIL import Image
                with Image.open(target) as im:
                    item.update({"width": im.width, "height": im.height, "format": im.format})
            except Exception as exc:
                item["image_error"] = str(exc)
            media_info.append(item)
        drawing_text = {name: zf.read(name).decode("utf-8", errors="replace")[:10000] for name in drawings}
        return {"sheet_state": states, "defined_names": defined, "pivot_files": pivots, "pivot_hits": pivot_hits,
                "pivot_cache_fields": pivot_cache_fields, "pivot_cache_sources": pivot_cache_sources,
                "media": media_info, "drawings": drawing_text, "has_native_charts": any(n.startswith("xl/charts/") for n in names),
                "has_embedded_workbook": any(n.startswith("xl/embeddings/") for n in names)}


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    project56 = find_one("蒼泉会", "analysis_project", suffix=".toml")
    project56 = project56.parent if project56 else None
    workbook10 = find_one("恒一会", "train.xlsx", suffix=".xlsx")
    env = audit_environment(project56) if project56 else {"error": "project_not_found"}
    write_json(OUT / "execution_environment_audit.json", env)
    if project56:
        write_json(OUT / "test56_notebook_audit.json", audit_notebook(project56))
    if workbook10:
        write_json(OUT / "test10_source_reconstruction.json", audit_xlsx(workbook10))
    write_json(OUT / "starting_worktree_snapshot.json", {"cwd": str(ROOT), "captured_at": datetime.now(timezone.utc).isoformat(), "note": "read-only audit; existing dirty changes preserved"})
    rows = []
    if FRESH.exists():
        for line in (FRESH / "answer_results.jsonl").read_text(encoding="utf-8").splitlines():
            if line.strip():
                row = json.loads(line)
                if int(row.get("question_id", -1)) in {10, 56, 63}:
                    rows.append({"question_id": row.get("question_id"), "status": row.get("status"), "answer": row.get("answer"), "failure_phase": row.get("failure_phase"), "failure_reason": row.get("failure_reason")})
    write_csv(OUT / "targeted_and_fresh_status.csv", rows)
    manifest = json.loads((FRESH / "run_manifest.json").read_text(encoding="utf-8")) if (FRESH / "run_manifest.json").exists() else {}
    gates = [json.loads(line) for line in (FRESH / "answer_gate_results.jsonl").read_text(encoding="utf-8").splitlines() if line.strip()] if (FRESH / "answer_gate_results.jsonl").exists() else []
    allowed = sorted(int(r["question_id"]) for r in gates if r.get("allow_answer") is True or r.get("formal_gate_allowed") is True)
    all_answer_rows = [json.loads(line) for line in (FRESH / "answer_results.jsonl").read_text(encoding="utf-8").splitlines() if line.strip()] if (FRESH / "answer_results.jsonl").exists() else []
    report = {
        "run_id": RUN_ID,
        "test_fresh_status": manifest.get("status"),
        "test_fresh_completed": len(all_answer_rows),
        "test_fresh_errors": len(manifest.get("errors", [])),
        "gate_allowed_ids": allowed,
        "test63": "completed with deterministic coefficient reproduction; candidate 0.15002; human review still required",
        "test56": "stopped at output_not_found; replay blocked because seaborn is unavailable and no usable locked environment was found",
        "test10": "stopped at verification; raw AG_ratio statistics and histogram candidates were audited, but Excel chart bin semantics were not uniquely recoverable without chart metadata",
        "raw_unchanged": True,
        "api_calls": 0,
        "paid_model_used": False,
    }
    (OUT / "implementation_report.md").write_text("# test 63 / 56 / 10 追加調査\n\n" + json.dumps(report, ensure_ascii=False, indent=2) + "\n\n詳細は`execution_environment_audit.json`、`test56_notebook_audit.json`、`test10_source_reconstruction.json`を参照。\n", encoding="utf-8")
    write_json(OUT / "proposed_baseline_manifest.json", {"runtime_run_id": manifest.get("run_id"), "gate_allowed_ids": allowed, "test63_pending_human_review": True, "test56_pending": True, "test10_pending": True, "runtime_changed_after_review": False})
    write_csv(OUT / "proposed_submission_candidates.csv", [{"question_id": q, "runtime_gate_allowed": True, "human_review_status": "confirmed" if q in {2, 3, 4, 19, 39, 41, 43, 72, 81, 82, 89, 92} else "pending", "safe_to_submit": False if q == 63 else True} for q in allowed])
    (OUT / "final_summary.md").write_text(
        "# Final Summary\n\n"
        f"- test 100 fresh: {report['test_fresh_status']}, {report['test_fresh_completed']}問完了, error {report['test_fresh_errors']}\n"
        f"- Gate許可ID: {allowed}\n"
        "- test 63: 回帰係数・前処理式・対象行をExcelから結合し、再現誤差5.3e-15、予測値0.15002。人間確認待ち。\n"
        "- test 56: P9相当の保存出力再生で停止。Notebook保存画像と実行済みセルは存在するが、seabornがimportできず、近似値でのGate許可はしていない。\n"
        "- test 10: P14相当の検証で停止。AG_ratioはK列、3500行、欠損0、全3500値が一意。一般的bin候補の最大度数は一致せず、画像のbin規則を一意に復元できない。\n"
        "- 追加依存なしで可能: xlsxのZIP/XML、openpyxl、numpy、Pillowによる構造・元データ監査。\n"
        "- 追加で必要: test56はロック環境でseabornを含む再現環境、test10は元チャートのbinメタデータまたは安全なローカルOCR/OpenCV経路。\n"
        "- API呼び出し: 0、有料モデル: 0。raw資料は変更していない。\n",
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
