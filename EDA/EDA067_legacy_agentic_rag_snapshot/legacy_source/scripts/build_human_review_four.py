from __future__ import annotations

import csv
import json
import pathlib
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parents[1]
OUT = ROOT / "data/output/human_review_four_candidates_v1/analysis"
FORMAL = ROOT / "data/output/question_file_operation_route_e2e_v1/e2e"
PREVIOUS = ROOT / "data/output/b2_autonomous_capability_expansion_test_final_v1"
QUESTIONS = next((p for p in (ROOT / "data/raw").rglob("questions_test.csv")), None)


def jsonl(path: pathlib.Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()] if path.exists() else []


def qtext() -> dict[str, str]:
    with QUESTIONS.open(encoding="utf-8-sig", newline="") as f:
        return {row["index"]: row["question"] for row in csv.DictReader(f)}


def style(cell) -> dict:
    fill = cell.fill
    color = fill.fgColor
    return {"fill_type": fill.fill_type, "fill_color_type": color.type, "fill_rgb": color.rgb, "fill_theme": color.theme, "number_format": cell.number_format, "bold": cell.font.bold, "italic": cell.font.italic, "formula": cell.data_type == "f", "has_comment": cell.comment is not None}


def cell_record(ws, row: int, col: int, data_ws=None) -> dict:
    cell = ws.cell(row, col)
    data_cell = data_ws.cell(row, col) if data_ws else cell
    return {"coordinate": cell.coordinate, "column": cell.column_letter, "row": row, "display_value": cell.value, "internal_value": data_cell.value, **style(cell)}


def context(ws, row: int, radius: int = 3) -> list[dict]:
    lo, hi = max(1, row - radius), min(ws.max_row, row + radius)
    return [{"row": r, "values": [ws.cell(r, c).value for c in range(1, ws.max_column + 1)], "styles": [style(ws.cell(r, c)) for c in range(1, ws.max_column + 1)]} for r in range(lo, hi + 1)]


def evidence_map(folder: pathlib.Path) -> dict[str, dict]:
    answers = {str(x["question_id"]): x for x in jsonl(folder / "answer_results.jsonl")}
    gates = {str(x["question_id"]): x for x in jsonl(folder / "answer_gate_results.jsonl")}
    return {qid: {"answer": answers.get(qid, {}), "gate": gates.get(qid, {})} for qid in {"2", "19", "82", "89"}}


def compact_row_values(ws, row: int) -> list[str]:
    return ["" if ws.cell(row, c).value is None else str(ws.cell(row, c).value) for c in range(1, ws.max_column + 1)]


def make_case(qid: str, source: pathlib.Path, rows: list[int], answer: str, target: str, conditions: str, logic: str, steps: str, classification: str, human_points: str, column: str, sheet_hint: str = "") -> tuple[dict, list[dict], str]:
    wb = load_workbook(source, data_only=False)
    wb_values = load_workbook(source, data_only=True)
    ws = wb[sheet_hint] if sheet_hint and sheet_hint in wb.sheetnames else wb.active
    ws_values = wb_values[ws.title]
    selected_cells = []
    contexts = []
    for row in rows:
        for col in range(1, ws.max_column + 1):
            selected_cells.append(cell_record(ws, row, col, ws_values))
        contexts.extend(context(ws, row))
    unique_context = {r["row"]: r for r in contexts}
    answer_cells = [cell_record(ws, row, ws[column + str(row)].column, ws_values) for row in rows]
    evidence = evidence_map(FORMAL).get(qid, {})
    if not evidence.get("answer"):
        evidence = evidence_map(PREVIOUS).get(qid, {})
    gate = evidence.get("gate", {})
    record = {"question_id": qid, "question_original": qtext()[qid], "answer_candidate": answer, "answer_format": "comma-separated list" if "," in answer else "text", "selected_source": str(source), "file_type": source.suffix.lstrip("."), "sheet_name": ws.title, "slide_number": "", "page_number": "", "table_name": "", "referenced_cells": ";".join(c["coordinate"] for c in answer_cells), "referenced_rows": ";".join(map(str, rows)), "referenced_columns": column, "referenced_ranges": f"A{min(rows)}:{get_column_letter(ws.max_column)}{max(rows)}", "question_target": target, "question_conditions": conditions, "condition_logic": logic, "same_row_required": "true", "extraction_method": "openpyxl raw workbook read; existing executor evidence cross-check", "calculation_or_selection_steps": steps, "source_values": json.dumps([c["display_value"] for c in answer_cells], ensure_ascii=False, default=str), "source_text": json.dumps([compact_row_values(ws, r) for r in rows], ensure_ascii=False, default=str), "neighboring_context": json.dumps(list(unique_context.values()), ensure_ascii=False, default=str), "evidence": json.dumps(evidence.get("answer", {}).get("evidence_locations", []), ensure_ascii=False, default=str), "verification_result": json.dumps(gate, ensure_ascii=False, default=str), "gate_result": gate.get("gate_status", "allowed"), "remaining_uncertainty": "色の意味、条件の漏れ、同一資料内の未検出行は人間確認が必要", "human_review_points": human_points, "classification": classification, "needs_human_review": "true", "safe_to_submit": "false"}
    md = f"""# test {qid}\n\n## 1. 質問文\n{qtext()[qid]}\n\n## 2. 現在の回答候補\n{answer}\n\n## 3. 使用資料\n`{source}`\nシート: `{ws.title}`\n\n## 4. 元資料の該当位置\n回答列: `{column}`、行: `{', '.join(map(str, rows))}`、セル: `{', '.join(c['coordinate'] for c in answer_cells)}`\n\n## 5. 該当データ\n```text\n{json.dumps([compact_row_values(ws, r) for r in rows], ensure_ascii=False, indent=2)}\n```\n\n## 6. 抽出・計算手順\n{steps}\n\n## 7. 条件ごとの成立確認\n- 対象: {target}\n- 条件: {conditions}\n- 論理: {logic}\n- raw上の回答列: {column}\n\n## 8. 漏れ・重複確認\n候補行は {len(rows)} 行。質問条件に一致する他行の有無と、色・日付・フェーズの解釈を人間が確認する。\n\n## 9. 曖昧な点\n{record['remaining_uncertainty']}\n\n## 10. 人間が最終確認する項目\n{human_points}\n\n## 11. H0〜H5の暫定分類\n**{classification}**\n\n既存Gate: `{record['gate_result']}`。`needs_human_review=true`、`safe_to_submit=false`を維持する。\n"""
    return record, selected_cells, md


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    qs = qtext()
    prior = evidence_map(PREVIOUS)
    cases = {}
    # 参照行・列は既存Evidenceから取得し、raw workbookで再確認する。
    specs = {
        "2": ("スケジュール_r2.xlsx", [2, 12, 21], "プロジェクトキックオフ実施, 中間報告会実施, 最終報告会実施", "タスク名", "オレンジ色でハイライトされた行", "color/style match", "色付き行を抽出し、同じ行のタスク名列Dを返す", "H1", "オレンジのRGB/テーマ色の意味、他の一致行、ヘッダー行を確認", "D", "スケジュール"),
        "82": ("スケジュール.xlsx", [3, 15, 17, 23, 25], "T02, T14, T16, T22, T24", "タスクID", "質問指定の色でハイライトされた行", "color/style match", "色付き行を抽出し、同じ行のタスクID列Cを返す", "H1", "色定義、5行以外の一致行、回答列Cの妥当性を確認", "C", "WBS"),
        "19": ("スケジュール_r2.xlsx", [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18], "T04, T05, T06, T07, T08, T09, T10, T11, T12, T13, T14, T15, T16, T17", "タスクID", "開始日または終了日が2025-08-11〜2025-09-09", "OR", "開始日と終了日を日付化し、範囲内を含む行をORで採用してタスクID列Cを返す", "H1", "境界日を含むか、開始/終了のOR、14件の漏れ・重複、質問の回答形式を確認", "C", "スケジュール"),
        "89": ("スケジュール.xlsx", [25], "最終報告・成果物提出・検収会", "タスク名", "フェーズNo. = 6、開始日が最大", "AND then MAX", "フェーズNo.6の行だけを残し、開始日の最大行のタスク名列を返す", "H1", "フェーズNo.6の全行、開始日最大値、同日同値の有無、タスク名列を確認", "D", "WBSタスク一覧"),
    }
    all_rows, all_cells, source_rows = [], [], []
    for qid, spec in specs.items():
        pattern, rows, answer, target, cond, logic, steps, cls, points, col, sheet = spec
        selected_paths = prior.get(qid, {}).get("answer", {}).get("selected_files", [])
        matches = [ROOT / p for p in selected_paths if (ROOT / p).exists()]
        # 既存runの選択パスを読み取り専用で辿る。Unicode正規化の影響を受ける名前検索は補助に留める。
        selected = matches[0] if matches else None
        if selected is None:
            matches = list((ROOT / "data/raw").rglob(pattern))
            selected = matches[0] if matches else None
        if selected is None:
            selected = matches[0] if matches else None
        if selected is None:
            raise FileNotFoundError(pattern)
        if qid == "89":
            # レビュー資料側も、固定行ではなく質問条件から最大開始日の行を再計算する。
            review_wb = load_workbook(selected, data_only=True)
            review_ws = review_wb[sheet] if sheet in review_wb.sheetnames else review_wb.active
            headers = {str(review_ws.cell(1, c).value): c for c in range(1, review_ws.max_column + 1)}
            phase_col = next((c for name, c in headers.items() if "フェーズNo" in name), None)
            date_col = next((c for name, c in headers.items() if "開始日" in name), None)
            if phase_col and date_col:
                candidates = []
                current_phase = None
                for r in range(2, review_ws.max_row + 1):
                    phase_value = review_ws.cell(r, phase_col).value
                    if phase_value not in (None, ""):
                        current_phase = str(phase_value).strip()
                    if current_phase in {"6", "6.0"} and review_ws.cell(r, date_col).value is not None:
                        candidates.append((review_ws.cell(r, date_col).value, r))
                if candidates:
                    latest = max(value for value, _ in candidates)
                    rows = [r for value, r in candidates if value == latest]
        record, cells, md = make_case(qid, selected, rows, answer, target, cond, logic, steps, cls, points, col, sheet)
        cases[qid] = record
        (OUT / f"test_{int(qid):03d}_review.md").write_text(md, encoding="utf-8")
        with (OUT / f"test_{int(qid):03d}_context.csv").open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["question_id", "row", "values", "styles"]); w.writeheader()
            wb = load_workbook(selected, data_only=False); ws = wb[sheet] if sheet in wb.sheetnames else wb.active
            for row in sorted(set(r for r in rows for r in range(max(1, r - 3), min(ws.max_row, r + 3) + 1))): w.writerow({"question_id": qid, "row": row, "values": json.dumps(compact_row_values(ws, row), ensure_ascii=False, default=str), "styles": json.dumps([style(ws.cell(row, c)) for c in range(1, ws.max_column + 1)], ensure_ascii=False, default=str)})
        all_rows.append(record); all_cells.extend({"question_id": qid, **cell} for cell in cells); source_rows.append({"question_id": qid, "source_path": str(selected), "sheet": record["sheet_name"], "cells": record["referenced_cells"], "ranges": record["referenced_ranges"]})
    fields = list(all_rows[0].keys())
    with (OUT / "human_review_questions.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(all_rows)
    with (OUT / "human_review_evidence.csv").open("w", encoding="utf-8", newline="") as f:
        fields = list(all_cells[0].keys()); w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(all_cells)
    with (OUT / "human_review_source_locations.csv").open("w", encoding="utf-8", newline="") as f:
        fields = list(source_rows[0].keys()); w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(source_rows)
    with (OUT / "unresolved_review_points.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["question_id", "review_point", "reason", "status"]); w.writeheader()
        for qid, row in cases.items():
            w.writerow({"question_id": qid, "review_point": row["human_review_points"], "reason": row["remaining_uncertainty"], "status": "human_review_required"})
    (OUT / "human_review_summary.md").write_text("# Human Review Summary\n\n対象: test 2, 19, 82, 89。4問とも回答候補・raw位置・周辺行を整理したが、Codex判定だけで提出可には変更していない。全問 `needs_human_review=true`、`safe_to_submit=false`。暫定分類は全てH1。\n", encoding="utf-8")
    (OUT / "next_unresolved_candidate_ranking.csv").write_text("question_id,reason,route_candidate,priority\n", encoding="utf-8")
    (OUT / "next_unresolved_candidate_summary.md").write_text("# Next Unresolved Candidates\n\n今回は4問の人間確認資料作成を優先した。追加候補は、資料・構造・Evidenceの一意性をこの監査だけでは確定できないため、順位付け対象なし。\n", encoding="utf-8")
    (OUT / "final_summary.md").write_text("# Final Summary\n\n4問のraw直接監査資料を生成した。runtimeコード、資料選択、Executor、Verification、Gateは変更していない。4問とも人間確認待ちを維持し、commit/push/PRは行っていない。\n", encoding="utf-8")


if __name__ == "__main__":
    main()
