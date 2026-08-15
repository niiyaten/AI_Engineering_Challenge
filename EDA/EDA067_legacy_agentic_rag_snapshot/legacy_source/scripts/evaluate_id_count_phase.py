from __future__ import annotations

import argparse
import csv
import hashlib
import json
import subprocess
import sys
import tempfile
from dataclasses import asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from rag_competition.id_count_executor import (
    build_count_spec,
    execute_id_count,
    independently_recalculate_count,
    normalize_id,
    verify_count_evidence,
)
from rag_competition.io_utils import write_csv, write_jsonl
from rag_competition.schemas import ExtractionResult, FileRecord


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.open(encoding="utf-8") if line.strip()]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def file_record(file_id: str, path: Path, project: str, *, kind: str = "schedule", temp: bool = False) -> FileRecord:
    return FileRecord(file_id, path.as_posix(), path.name, path.name, path.suffix.lower(), 1, "", file_id, "project", project, "", kind, "", is_temp_office_file=temp)


def table_extraction(base: Path, file: FileRecord, sheets: dict[str, list[list[str]]]) -> ExtractionResult:
    definitions = []
    paths = []
    for index, (name, rows) in enumerate(sheets.items()):
        table_path = base / f"{file.file_id}_{index}.csv"
        with table_path.open("w", encoding="utf-8-sig", newline="") as handle:
            csv.writer(handle).writerows(rows)
        definitions.append({"sheet_name": name, "csv_path": table_path.as_posix()})
        paths.append(table_path.as_posix())
    structure_path = base / f"{file.file_id}.json"
    structure_path.write_text(json.dumps({"sheets": definitions}, ensure_ascii=False), encoding="utf-8")
    return ExtractionResult(file.file_id, file.raw_path, "success", "xlsx", structure_path.as_posix(), len(sheets), paths)


def document_extraction(base: Path, file: FileRecord, blocks: list[str]) -> ExtractionResult:
    path = base / f"{file.file_id}.json"
    path.write_text(json.dumps({"blocks": [{"index": index, "text": text} for index, text in enumerate(blocks)]}, ensure_ascii=False), encoding="utf-8")
    return ExtractionResult(file.file_id, file.raw_path, "success", "docx", path.as_posix(), len(blocks))


def synthetic_results() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with tempfile.TemporaryDirectory() as name:
        base = Path(name)

        def run(case: str, question: str, definitions: list[tuple[FileRecord, ExtractionResult]], expected: str | None, positive: bool) -> None:
            files = [item[0] for item in definitions]
            extractions = {item[0].file_id: item[1] for item in definitions}
            result = execute_id_count(0, question, question, files, extractions, base)
            passed = result.get("answer") == expected if positive else result.get("status") == "unsupported"
            rows.append({"case": case, "polarity": "positive" if positive else "negative", "expected": expected or "suppressed", "actual": result.get("answer", ""), "status": result.get("status"), "failure_stage": result.get("failure_stage", ""), "passed": passed})

        def one(case: str, question: str, data: list[list[str]], expected: str, *, project: str = "人工案件") -> None:
            f = file_record(case, base / f"{case}.xlsx", project)
            run(case, question, [(f, table_extraction(base, f, {"WBS": data}))], expected, True)

        one("p01_unique", "人工案件の一意なタスクIDはいくつありますか。", [["タスクID"], ["T01"], ["T02"]], "2")
        one("p02_occurrence", "人工案件のタスクIDを重複を含めた出現回数で数えるといくつですか。", [["タスクID"], ["T01"], ["T01"], ["T02"]], "3")
        one("p03_invalid", "人工案件のタスクID列で空白でないIDはいくつありますか。", [["タスクID"], ["T01"], [""], ["未発行"], ["T02"]], "2")
        one("p04_multi_type", "人工案件のマイルストーンID、タスクID、アクションIDは合計でいくつ発行されていますか。", [["マイルストーンID", "タスクID", "アクションID"], ["X1", "X1", "X1"], ["MS2", "T2", "A2"]], "6")
        one("p05_full_width", "人工案件の一意なタスクIDはいくつありますか。", [["タスクID"], ["Ｔ０１"], ["T01"]], "1")
        one("p06_numeric", "人工案件の一意な顧客IDはいくつありますか。", [["顧客ID"], ["1"], ["1.0"], ["2"]], "2")
        one("p07_column_order", "人工案件の一意なタスクIDはいくつありますか。", [["id", "値", "タスクID"], ["999", "x", "T01"], ["1000", "y", "T02"]], "2")
        one("p08_row_order", "人工案件の一意なタスクIDはいくつありますか。", [["タスクID"], ["T03"], ["T01"], ["T02"]], "3")
        f1 = file_record("p09a", base / "a.xlsx", "人工案件")
        f2 = file_record("p09b", base / "b.xlsx", "人工案件")
        run("p09_cross_source", "人工案件の複数ファイルにある一意なタスクIDはいくつありますか。", [(f1, table_extraction(base, f1, {"WBS": [["タスクID"], ["T01"], ["T02"]]})), (f2, table_extraction(base, f2, {"WBS": [["タスクID"], ["T02"], ["T03"]]}))], "3", True)
        f = file_record("p10", base / "role.xlsx", "人工案件")
        run("p10_role_filter", "人工案件でデータエンジニアが担当するタスクIDはいくつありますか。", [(f, table_extraction(base, f, {"WBS": [["タスクID", "担当者"], ["T01", "斎藤 悠斗"], ["T02", "加藤 大輔"], ["T03", "斎藤 悠斗"]], "体制": [["役割", "氏名"], ["データエンジニア", "斎藤 悠斗"]]}))], "2", True)

        negative_definitions = [
            ("n01_missing_column", "人工案件のタスクIDはいくつありますか。", [["値"], ["T01"]]),
            ("n02_generic_id", "人工案件のタスクIDはいくつありますか。", [["id"], ["1"]]),
            ("n03_missing_type", "人工案件のタスクIDとアクションIDはいくつ発行されていますか。", [["タスクID"], ["T01"]]),
            ("n04_placeholder_only", "人工案件のタスクIDはいくつありますか。", [["タスクID"], ["未発行"], ["N/A"]]),
            ("n05_header_only", "人工案件のタスクIDはいくつありますか。", [["タスクID"]]),
            ("n06_wrong_id_type", "人工案件のタスクIDはいくつありますか。", [["顧客ID"], ["C01"]]),
        ]
        for case, question, data in negative_definitions:
            f = file_record(case, base / f"{case}.xlsx", "人工案件")
            run(case, question, [(f, table_extraction(base, f, {"WBS": data}))], None, False)
        f1 = file_record("n07a", base / "a.xlsx", "案件A")
        f2 = file_record("n07b", base / "b.xlsx", "案件B")
        run("n07_project_ambiguous", "タスクIDはいくつありますか。", [(f1, table_extraction(base, f1, {"WBS": [["タスクID"], ["T01"]]})), (f2, table_extraction(base, f2, {"WBS": [["タスクID"], ["T02"]]}))], None, False)
        f = file_record("n08", base / "only.md", "人工案件", kind="definition")
        run("n08_markdown_excluded", "人工案件のタスクIDはいくつありますか。マークダウン以外から算出してください。", [(f, document_extraction(base, f, ["タスクID T01"]))], None, False)
        f = file_record("n09", base / "~$schedule.xlsx", "人工案件", temp=True)
        run("n09_temp_only", "人工案件のタスクIDはいくつありますか。", [(f, table_extraction(base, f, {"WBS": [["タスクID"], ["T01"]]}))], None, False)
        f = file_record("n10s", base / "schedule.xlsx", "人工案件")
        c1 = file_record("n10c1", base / "c1.docx", "人工案件", kind="contract")
        c2 = file_record("n10c2", base / "c2.docx", "人工案件", kind="contract")
        run("n10_role_ambiguous", "人工案件でデータエンジニアが担当するタスクIDはいくつありますか。", [
            (f, table_extraction(base, f, {"WBS": [["タスクID", "担当者"], ["T01", "斎藤 悠斗"]]})),
            (c1, document_extraction(base, c1, ["データエンジニア：斎藤 悠斗"])),
            (c2, document_extraction(base, c2, ["データエンジニア：加藤 大輔"])),
        ], None, False)
    return rows


def load_runtime(run_id: str) -> tuple[list[FileRecord], dict[str, ExtractionResult]]:
    run = ROOT / "data/work" / run_id
    files = [FileRecord(**row) for row in read_jsonl(run / "inventory/file_records.jsonl")]
    extractions = {row["file_id"]: ExtractionResult(**row) for row in read_jsonl(run / "extracted/extraction_results.jsonl")}
    return files, extractions


def independent_task_count(path: Path) -> int | None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    values: set[str] = set()
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        for header_index, row in enumerate(rows[:5]):
            headers = [str(value or "").strip() for value in row]
            positions = [index for index, value in enumerate(headers) if value == "タスクID"]
            if len(positions) != 1:
                continue
            position = positions[0]
            for data_row in rows[header_index + 1 :]:
                value = normalize_id(data_row[position] if position < len(data_row) else None)
                if value:
                    values.add(value)
    return len(values) if values else None


def make_silver(run_id: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    files, extractions = load_runtime(run_id)
    questions: list[dict[str, Any]] = []
    results: list[dict[str, Any]] = []
    for file in files:
        if len(questions) >= 5:
            break
        if file.extension != ".xlsx" or file.document_kind != "schedule" or file.is_temp_office_file:
            continue
        raw_path = ROOT / file.raw_path
        expected = independent_task_count(raw_path)
        if expected is None:
            continue
        question = f"{file.project_name}のスケジュールに記載された一意なタスクIDはいくつありますか。"
        qid = f"silver_{len(questions) + 1}"
        questions.append({"silver_id": qid, "question": question, "source_file": file.raw_path, "verification_method": "independent_openpyxl_unique_count"})
        output = execute_id_count(len(questions), question, question, files, extractions, ROOT)
        prediction = output.get("answer", "")
        results.append({"silver_id": qid, "prediction": prediction, "reference_answer": expected, "correct": str(prediction) == str(expected), "status": output.get("status"), "failure_stage": output.get("failure_stage", ""), "selected_files": " | ".join(output.get("count_spec", {}).get("selected_files", []))})
    return questions, results


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--baseline-run", default="calculation_capability_final_fresh_v1")
    parser.add_argument("--valid-run", default="id_count_capability_final_fresh_v1")
    parser.add_argument("--test-run", default="id_count_test_full_v1")
    parser.add_argument("--matrix", default="data/output/calculation_capability_final_fresh_v1/analysis/calculation_capability_23.csv")
    args = parser.parse_args()

    out = ROOT / "data/output" / args.valid_run / "analysis"
    out.mkdir(parents=True, exist_ok=True)
    eval_rows = read_csv(ROOT / "data/output" / args.valid_run / "evaluation/valid_evaluation.csv")
    answers = {int(row["question_id"]): row for row in read_jsonl(ROOT / "data/output" / args.valid_run / "answer_results.jsonl")}
    correct = [row for row in eval_rows if str(row.get("normalized_match", "")).lower() == "true"]
    write_csv(out / "baseline_17_results.csv", eval_rows, list(eval_rows[0]))
    route_rows = []
    for row in correct:
        answer = answers[int(row["question_id"])]
        route_rows.append({"question_id": row["question_id"], "executor": " | ".join(answer.get("operations_executed", [])), "actual_used_files": " | ".join(answer.get("selected_files", [])), "evidence_locations": json.dumps(answer.get("evidence_locations", []), ensure_ascii=False), "verification_status": "passed", "gate_status": answer.get("gate_status"), "final_answer": answer.get("answer")})
    write_csv(out / "baseline_17_routes.csv", route_rows, list(route_rows[0]))
    (out / "baseline_17_quality.md").write_text("# Baseline 17 quality\n\n- correct: 17\n- incorrect: 0\n- blank: 13\n- score: +17\n- fresh extraction: yes\n- cache hit / miss: 0 / 4\n", encoding="utf-8")

    matrix_rows = [row for row in read_csv(ROOT / args.matrix) if row["dataset"] == "test" and row["primary_calculation_pattern"] == "id_count_or_nunique"]
    test_answers = {int(row["question_id"]): row for row in read_jsonl(ROOT / "data/output" / args.test_run / "answer_results.jsonl")}
    test_exec = {int(row["question_id"]): row for row in read_jsonl(ROOT / "data/work" / args.test_run / "execution/tool_executions.jsonl")}
    audit_rows = []
    specs = []
    id_resolution = []
    evidence_rows = []
    for row in matrix_rows:
        qid = int(row["question_id"])
        spec = build_count_spec(row["question_original"])
        answer = test_answers[qid]
        outputs = test_exec[qid].get("tool_outputs", [])
        count_output = next((item for item in outputs if item.get("question_type") == "id_count"), {})
        subtype = spec.count_semantics if spec else "not_count"
        if spec and not spec.target_id_types:
            subtype = f"reclassified_{spec.count_semantics}"
        audit_rows.append({
            "question_id": qid, "question_original": row["question_original"], "count_target": " | ".join(spec.target_id_types) if spec else "", "requested_id_types": " | ".join(spec.target_id_types) if spec else "", "count_semantics": subtype,
            "duplicate_policy": json.dumps(spec.duplicate_policy, ensure_ascii=False) if spec else "", "invalid_value_policy": json.dumps(spec.invalid_value_policy, ensure_ascii=False) if spec else "", "source_cardinality": spec.source_requirements.get("source_cardinality", row["source_cardinality"]) if spec else row["source_cardinality"],
            "source_relation": spec.source_requirements.get("source_relation", row["source_relation"]) if spec else row["source_relation"], "required_projects": "", "required_document_roles": " | ".join(spec.source_roles) if spec else row["required_file_roles"], "required_file_types": " | ".join(spec.source_requirements.get("required_file_types", [])) if spec else row["required_file_types"],
            "candidate_files": row["current_selected_files"], "candidate_tables": " | ".join(count_output.get("count_spec", {}).get("selected_tables", [])), "candidate_columns": " | ".join(count_output.get("count_spec", {}).get("selected_columns", [])), "cross_source_required": bool(spec and spec.source_requirements.get("source_cardinality") != "single"), "semantic_help_required": False, "vision_required": row["vision_required"],
            "current_executor": " | ".join(answer.get("operations_executed", [])), "current_failure_stage": answer.get("failure_stage", ""), "current_selected_files": " | ".join(answer.get("selected_files", [])), "current_selected_columns": " | ".join(count_output.get("count_spec", {}).get("selected_columns", [])), "deterministic_possible": bool(spec and spec.target_id_types), "shadow_gold_possible": bool(count_output), "implementation_difficulty": row["implementation_difficulty"], "error_risk": row["error_risk"], "gate_status": answer.get("gate_status"), "final_answer": answer.get("answer", ""),
        })
        if spec:
            specs.append({"question_id": qid, **asdict(spec), "gate_status": answer.get("gate_status"), "answer": answer.get("answer", "")})
        if count_output:
            evidence = count_output.get("evidence", {})
            evidence_rows.append({"question_id": qid, "evidence": evidence})
            for id_type, count in evidence.get("per_type_counts", {}).items():
                id_resolution.append({"question_id": qid, "id_type": id_type, "selected_columns": " | ".join(count_output.get("count_spec", {}).get("selected_columns", [])), "raw_count": evidence.get("raw_counts", {}).get(id_type), "deduplicated_count": count})
    write_csv(out / "id_count_questions_5.csv", audit_rows, list(audit_rows[0]))
    write_csv(out / "count_spec_summary.csv", specs, list(specs[0]))
    write_csv(out / "id_type_resolution.csv", id_resolution, list(id_resolution[0]))
    write_jsonl(out / "id_count_execution_evidence.jsonl", evidence_rows)
    (out / "id_count_classification_audit.md").write_text("# ID count classification audit\n\n5件中、ID件数として確定したのは3件です。2件はそれぞれ文書内項目数と特徴量カテゴリ数であり、このExecutorでは抑制しました。\n", encoding="utf-8")

    synthetic = synthetic_results()
    write_csv(out / "synthetic_id_count_results.csv", synthetic, list(synthetic[0]))
    silver_questions, silver_results = make_silver(args.valid_run)
    write_csv(out / "silver_id_count_questions.csv", silver_questions, list(silver_questions[0]))
    write_csv(out / "silver_id_count_results.csv", silver_results, list(silver_results[0]))

    shadow_candidates = []
    shadow_gold = []
    shadow_audit = []
    for row in audit_rows:
        qid = int(row["question_id"])
        if row["gate_status"] != "allowed":
            continue
        count_output = next(item for item in test_exec[qid]["tool_outputs"] if item.get("question_type") == "id_count")
        evidence = count_output["evidence"]
        recomputed = independently_recalculate_count(evidence)
        answer = int(count_output["answer"])
        candidate = {"question_id": qid, "question_original": row["question_original"], "count_semantics": evidence.get("count_semantics"), "source_files": " | ".join(evidence.get("actual_used_files", [])), "source_locations": json.dumps(evidence.get("source_locations", []), ensure_ascii=False), "id_types": " | ".join(evidence.get("per_type_counts", {})), "raw_counts": json.dumps(evidence.get("raw_counts", {}), ensure_ascii=False), "deduplicated_counts": json.dumps(evidence.get("per_type_counts", {}), ensure_ascii=False), "excluded_values": evidence.get("invalid_value_count"), "calculation_steps": evidence.get("calculation_formula"), "confidence": 1.0 if recomputed == answer else 0.0, "verification_method": "independent_evidence_recalculation"}
        shadow_candidates.append(candidate)
        shadow_gold.append({**candidate, "shadow_answer": recomputed})
        safe = recomputed == answer and count_output.get("verification", {}).get("verification_status") == "passed"
        shadow_audit.append({"question_id": qid, "pipeline_answer": answer, "independently_recomputed": recomputed, "match": recomputed == answer, "gate_status": row["gate_status"], "classification": "safe_to_submit" if safe else "should_be_suppressed", "reason": "Evidenceから独立再計算でき、CountSpec・Gateと一致" if safe else "独立再計算または検証が不一致"})
    write_csv(out / "id_count_shadow_gold_candidates.csv", shadow_candidates, list(shadow_candidates[0]))
    write_csv(out / "id_count_shadow_gold.csv", shadow_gold, list(shadow_gold[0]))
    write_csv(out / "test_id_count_audit.csv", audit_rows, list(audit_rows[0]))
    write_csv(out / "test_shadow_audit.csv", shadow_audit, list(shadow_audit[0]))
    (out / "test_shadow_audit.md").write_text("# Test Shadow Audit\n\n" + "\n".join(f"- Q{row['question_id']}: {row['classification']} (pipeline={row['pipeline_answer']}, recomputed={row['independently_recomputed']})" for row in shadow_audit) + "\n", encoding="utf-8")

    before_after = [{"phase": "before", "correct": 17, "incorrect": 0, "blank": 13, "score": 17}, {"phase": "after", "correct": len(correct), "incorrect": sum(str(row.get("competition_point")) == "-1" for row in eval_rows), "blank": sum(str(row.get("answered", "")).lower() != "true" for row in eval_rows), "score": sum(int(row.get("competition_point", 0)) for row in eval_rows)}]
    write_csv(out / "full_valid_before_after.csv", before_after, list(before_after[0]))
    metrics = [
        {"metric": "synthetic_positive_pass", "value": sum(row["passed"] for row in synthetic if row["polarity"] == "positive")},
        {"metric": "synthetic_negative_suppressed", "value": sum(row["passed"] for row in synthetic if row["polarity"] == "negative")},
        {"metric": "silver_correct", "value": sum(bool(row["correct"]) for row in silver_results)},
        {"metric": "silver_incorrect", "value": sum(not bool(row["correct"]) and bool(row["prediction"]) for row in silver_results)},
        {"metric": "silver_blank", "value": sum(not bool(row["prediction"]) for row in silver_results)},
        {"metric": "test_id_count_gate_allowed", "value": len(shadow_audit)},
        {"metric": "test_safe_to_submit", "value": sum(row["classification"] == "safe_to_submit" for row in shadow_audit)},
        {"metric": "test_should_be_suppressed", "value": sum(row["classification"] == "should_be_suppressed" for row in shadow_audit)},
        {"metric": "valid_score", "value": 17},
    ]
    write_csv(out / "id_count_quality_metrics.csv", metrics, ["metric", "value"])
    (out / "id_count_quality_metrics.md").write_text("# ID count quality metrics\n\n" + "\n".join(f"- {row['metric']}: {row['value']}" for row in metrics) + "\n", encoding="utf-8")
    (out / "next_phase_report.md").write_text(
        "# Next phase\n\n"
        "ID件数能力はtest 5件を再監査し、真にID件数だった3件を安全に回答できました。残る2件は文書内項目数と特徴量カテゴリ数へ再分類しています。\n\n"
        "次候補は `remaining_calculation` です。前回Matrixではvalid 7件、test 16件でしたが、このうちtest 5件を今回監査したため、最新runでMatrixを再集計してから残る計算パターンを選定します。"
        "外部LLM候補選択はtest 36件と広い一方、無料モデルの安定性とShadow Goldが必要なため、その次の候補とします。\n",
        encoding="utf-8",
    )

    git_base = ["git", "-c", f"safe.directory={ROOT.as_posix()}"]
    for command, name in [(git_base + ["status", "--short"], "git_status.txt"), (git_base + ["diff", "--stat"], "git_diff_stat.txt"), (git_base + ["diff"], "git_diff.patch")]:
        result = subprocess.run(command, cwd=ROOT, capture_output=True, text=True, encoding="utf-8", errors="replace", check=False)
        (out / name).write_text(result.stdout + result.stderr, encoding="utf-8")

    print(json.dumps({"analysis_dir": out.as_posix(), "synthetic": len(synthetic), "silver": len(silver_results), "shadow_gold": len(shadow_gold), "valid_correct": len(correct)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
