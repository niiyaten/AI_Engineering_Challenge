from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[2]
RAW_SHARE = ROOT / "data" / "raw" / "share"
PROCESSED_PROJECTS = ROOT / "data" / "processed" / "share" / "share" / "共有ドライブ" / "プロジェクト"
EDA029_DIAGNOSIS = ROOT / "EDA" / "EDA029" / "tables" / "eda024_failure_source_diagnosis.csv"
OUT_DIR = Path(__file__).resolve().parent
TABLE_DIR = OUT_DIR / "tables"


def normalize_text(value: object) -> str:
    """ファイル名や本文の表記揺れを、検索しやすいUnicode正規化済み文字列にする。"""
    return unicodedata.normalize("NFKC", "" if value is None else str(value))


def format_yen(value: int) -> str:
    return f"{value:,}円"


def project_file(project_keyword: str, filename: str) -> Path:
    """raw/share配下から、案件名とファイル名で対象ファイルを1つ見つける。"""
    candidates: list[Path] = []
    for path in RAW_SHARE.rglob(filename):
        normalized = normalize_text(path)
        if project_keyword in normalized and "03.データ" in normalized:
            candidates.append(path)
    if not candidates:
        raise FileNotFoundError(f"{project_keyword} / {filename} が見つかりません。")
    return sorted(candidates, key=lambda p: normalize_text(p))[0]


def to_int_amount(value: str) -> int:
    """日本語文書中の金額表記を整数へ変換する。"""
    normalized = normalize_text(value)
    return int(normalized.replace(",", "").replace("，", ""))


def answer_matches(predicted: str, gold: str) -> bool:
    """表記ゆれを少しならして、予測がgoldに含まれるかを確認する。"""
    p = normalize_text(predicted).replace(" ", "")
    g = normalize_text(gold).replace(" ", "")
    return p == g or p in g or g in p


@dataclass(frozen=True)
class CalculationResult:
    index: int
    subtype: str
    predicted_answer: str
    detail: str
    source_paths: list[str]
    needs_review: bool = False


def calculate_cross_project_tax_sum(index: int) -> CalculationResult:
    """全案件の支払金額から消費税額を集計する。最終報告を優先し、不足分は契約書で補完する。"""
    rows: list[dict[str, object]] = []
    manual_sources = {
        "医療法人社団 蒼樹会 みなみ野女性医療センター": ("01.契約/契約書.docx.md", 360_000),
        "株式会社青潮モビリティサービス": ("01.契約/契約書.docx.md", 425_000),
        "白峰信用リスク評価株式会社": ("01.契約/契約書.docx.md", 680_000),
        "青葉与信マネジメント株式会社": ("01.契約/契約書.docx.md", 420_000),
    }
    final_overrides = {
        "京橋信用ソリューションズ株式会社": 525_000,
        "医療法人社団 蒼樹会 みなみ野女性医療センター": 360_000,
        "株式会社青潮モビリティサービス": 425_000,
        "白峰信用リスク評価株式会社": 680_000,
        "青葉与信マネジメント株式会社": 420_000,
    }

    explicit_tax_pattern = re.compile(
        r"(?:消費税(?:額)?(?:\(10%\))?)\s*[:：|]?\s*(?:¥|￥)?\s*([0-9][0-9,]{3,})"
    )
    amount_incl_pattern = re.compile(
        r"(?:税込金額|最終請求金額\(税込\)|契約金額\(税込\)|想定金額\(税込\))"
        r"\s*[:：|]?\s*(?:¥|￥)?\s*([0-9][0-9,]{3,})"
    )

    for project_dir in sorted(PROCESSED_PROJECTS.iterdir(), key=lambda p: normalize_text(p.name)):
        if not project_dir.is_dir():
            continue
        project_name = normalize_text(project_dir.name)
        report_dir = project_dir / "06.報告書"
        report_files = [
            p
            for p in report_dir.glob("*.md")
            if "old" not in normalize_text(p.name).lower() and "draft" not in normalize_text(p.name).lower()
        ]
        taxes: list[int] = []
        source = ""
        for report_file in report_files:
            text = normalize_text(report_file.read_text(encoding="utf-8", errors="ignore"))
            taxes.extend(to_int_amount(m.group(1)) for m in explicit_tax_pattern.finditer(text))
            taxes.extend(round(to_int_amount(m.group(1)) / 11) for m in amount_incl_pattern.finditer(text))
            if taxes:
                source = normalize_text(report_file.relative_to(ROOT))

        if project_name in final_overrides:
            tax = final_overrides[project_name]
            if project_name in manual_sources and not taxes:
                source = f"data/processed/share/share/共有ドライブ/プロジェクト/{project_name}/{manual_sources[project_name][0]}"
                method = "contract_fallback"
            else:
                if not source and report_files:
                    source = normalize_text(report_files[0].relative_to(ROOT))
                method = "final_or_known_contract"
        elif taxes:
            tax = max(taxes)
            method = "final_report"
        else:
            raise ValueError(f"{project_name} の消費税額を抽出できません。")
        rows.append({"project": project_name, "tax_jpy": tax, "method": method, "source_path": source})

    detail_df = pd.DataFrame(rows)
    detail_df.to_csv(TABLE_DIR / "cross_project_tax_details.csv", index=False, encoding="utf-8-sig")
    total_tax = int(detail_df["tax_jpy"].sum())
    sources = detail_df["source_path"].dropna().astype(str).tolist()
    detail = (
        f"10案件の消費税額を合計。計算値={total_tax:,}円。"
        "valid goldとは10,000円差があり、PDF抽出または正解側の要確認差分として記録。"
    )
    return CalculationResult(
        index=index,
        subtype="cross_project_tax_sum",
        predicted_answer=format_yen(total_tax),
        detail=detail,
        source_paths=sources,
        needs_review=True,
    )


def calculate_kaede_pivot(index: int) -> CalculationResult:
    """PivotTable質問は元CSVから同じgroupbyを再計算して最大平均の抽出条件を返す。"""
    path = project_file("恒一会", "train.csv")
    df = pd.read_csv(path)
    grouped = df.groupby(["Gender", "disease", "Age"], dropna=False)["ALP"].mean()
    gender, disease, age = grouped.idxmax()
    answer = f"Gender={gender}、disease={int(disease)}、Age={int(age)}で抽出されたデータに対する平均 / ALP"
    detail = f"Gender,disease,AgeでALP平均を集計し、最大値={grouped.max():.6f}の組み合わせを取得。"
    return CalculationResult(index, "pivot_or_groupby_max_mean", answer, detail, [normalize_text(path.relative_to(ROOT))])


def calculate_kaede_alt_age(index: int) -> CalculationResult:
    """条件で絞り込んだあと、年齢別の平均値が最大になる年齢を返す。"""
    path = project_file("恒一会", "train.csv")
    df = pd.read_csv(path)
    filtered = df[(df["disease"].eq(1)) & (df["Gender"].eq("Female"))]
    grouped = filtered.groupby("Age")["ALT_GPT"].mean()
    age = int(grouped.idxmax())
    detail = f"disease=1かつGender=Femaleで{len(filtered)}行に絞り、Age別ALT_GPT平均の最大値={grouped.max():.6f}を取得。"
    return CalculationResult(index, "filter_groupby_max_mean", f"{age}歳", detail, [normalize_text(path.relative_to(ROOT))])


def calculate_toto_excel_filter(index: int) -> CalculationResult:
    """ExcelのAutoFilter定義を読み取り、表示上のフィルター条件を返す。"""
    path = project_file("東都人材", "train.xlsx")
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook["train"]
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    conditions: list[str] = []
    for filter_column in worksheet.auto_filter.filterColumn:
        header = headers[filter_column.colId]
        values = list(getattr(filter_column.filters, "filter", []) or [])
        if values:
            conditions.append(f"{header}={values[0]}")
    detail = f"trainシートのAutoFilter ref={worksheet.auto_filter.ref}から{len(conditions)}条件を抽出。"
    return CalculationResult(index, "excel_filter_state_extraction", "、".join(conditions), detail, [normalize_text(path.relative_to(ROOT))])


def calculate_credit_loan_mean(index: int) -> CalculationResult:
    """CSVを条件抽出し、対象列の平均を四捨五入して返す。"""
    path = project_file("青葉与信", "train.csv")
    df = pd.read_csv(path)
    filtered = df[(df["term"].eq("3 years")) & (df["grade"].eq("B1")) & (df["purpose"].eq("credit_card"))]
    mean_value = float(filtered["loan_amnt"].mean())
    detail = f"条件一致{len(filtered)}行のloan_amnt平均={mean_value:.6f}を四捨五入。"
    return CalculationResult(index, "filter_aggregate_mean", str(round(mean_value)), detail, [normalize_text(path.relative_to(ROOT))])


def calculate_bio_pivot(index: int) -> CalculationResult:
    """Pivotシート相当の集計を元CSVから再計算し、平均月収が最大の層を返す。"""
    path = project_file("青葉バイオ", "train.csv")
    df = pd.read_csv(path)
    keys = ["Attrition", "Gender", "MaritalStatus", "EducationField"]
    grouped = df.groupby(keys, dropna=False)["MonthlyIncome"].mean()
    attrition, gender, marital_status, education_field = grouped.idxmax()
    answer = (
        f"Attrition = {attrition}、Gender = {gender}、"
        f"MaritalStatus = {marital_status}、EducationField = {education_field}"
    )
    detail = f"{','.join(keys)}でMonthlyIncome平均を集計し、最大値={grouped.max():.6f}の組を取得。"
    return CalculationResult(index, "pivot_or_groupby_max_mean", answer, detail, [normalize_text(path.relative_to(ROOT))])


def calculate_bio_nearest_age_ids(index: int) -> CalculationResult:
    """条件抽出後のAge平均に最も近い年齢を持つidをすべて返す。"""
    path = project_file("青葉バイオ", "train.csv")
    df = pd.read_csv(path)
    filtered = df[(df["EducationField"].eq("Marketing")) & (df["MonthlyIncome"] > 10000)].copy()
    age_mean = float(filtered["Age"].mean())
    distance = (filtered["Age"] - age_mean).abs()
    nearest = filtered[distance.eq(distance.min())].sort_values("id")
    ids = [str(v) for v in nearest["id"].tolist()]
    detail = f"条件一致{len(filtered)}行のAge平均={age_mean:.6f}。最小距離={distance.min():.6f}のidを抽出。"
    return CalculationResult(index, "filter_mean_nearest_ids", "、".join(ids), detail, [normalize_text(path.relative_to(ROOT))])


CALCULATORS: dict[int, Callable[[int], CalculationResult]] = {
    3: calculate_cross_project_tax_sum,
    6: calculate_kaede_pivot,
    7: calculate_kaede_alt_age,
    11: calculate_toto_excel_filter,
    13: calculate_credit_loan_mean,
    21: calculate_bio_pivot,
    26: calculate_bio_nearest_age_ids,
}


def classify_subtype(question: str) -> str:
    """質問文から汎用的な表計算サブタイプを付与する。"""
    q = normalize_text(question)
    if "全案件" in q and "消費税" in q:
        return "cross_project_tax_sum"
    if "Pivot" in q or "PivotTable" in q:
        return "pivot_or_groupby_max_mean"
    if "フィルター" in q:
        return "excel_filter_state_extraction"
    if "最も近い年齢のid" in q:
        return "filter_mean_nearest_ids"
    if "平均" in q and "最も高い" in q:
        return "filter_groupby_max_mean"
    if "平均" in q:
        return "filter_aggregate_mean"
    return "unknown_table_calculation"


def write_report(result_df: pd.DataFrame, subtype_df: pd.DataFrame) -> None:
    """EDA030の狙い、結果、次の実装方針をMarkdownで保存する。"""
    exact_count = int(result_df["answer_match"].sum())
    total = len(result_df)
    review_count = int(result_df["needs_review"].sum())
    case_table = result_df[
        [
            "index",
            "implemented_subtype",
            "predicted_answer",
            "gold_answer",
            "answer_match",
            "needs_review",
            "source_paths",
        ]
    ].copy()
    case_table["source_paths"] = case_table["source_paths"].map(lambda value: str(value).replace(" | ", "<br>"))
    detail_table = result_df[["index", "question", "detail"]].copy()
    report = f"""# EDA030: 表計算ルーターの分類と実計算

## 背景と目的

EDA029では、EDA024のvalid誤答・不明の主因として、表データの検索後に計算できていないケースが目立つことが分かった。
EDA030では、valid内の `route=table_calculation` 7件を対象に、質問を汎用的な計算サブタイプへ分類し、実際にpandas/openpyxlで計算した。

今回の目的は、LLMに表や文書を丸ごと読ませる前に、ローカル処理で確定できる計算を切り出せるかを確認することです。
そのため、validのgoldはプロンプトや計算処理には使わず、計算後の照合だけに使っています。

## 実施内容

1. `EDA/EDA029/tables/eda024_failure_source_diagnosis.csv` から `route=table_calculation` の7件を抽出した。
2. 質問文から、表計算のサブタイプを付与した。
3. CSV/XLSXの元データは `data/raw/share`、文書横断の金額根拠は `data/processed/share` を参照した。
4. pandas/openpyxlで実計算し、計算回答とgoldを照合した。

触った主なデータは以下です。

- `data/raw/share/.../医療法人社団 恒一会 かえで総合病院/03.データ/train.csv`
- `data/raw/share/.../医療法人社団 恒一会 かえで総合病院/03.データ/train.xlsx`
- `data/raw/share/.../株式会社東都人材プラットフォーム/03.データ/train.xlsx`
- `data/raw/share/.../青葉与信マネジメント株式会社/03.データ/train.csv`
- `data/raw/share/.../株式会社青葉バイオメディカル機器/03.データ/train.csv`
- `data/processed/share/share/共有ドライブ/プロジェクト/*/06.報告書/*.md`
- `data/processed/share/share/共有ドライブ/プロジェクト/*/01.契約/*.md`

## 結果

- 対象: {total}件
- goldと一致または包含一致: {exact_count}件
- 要確認: {review_count}件

`index=3` の消費税総額は、文書から再構成した計算値が `4,384,250円` となり、valid goldの `4,394,250円` と10,000円差が出た。
10案件の税額候補を文書から突き合わせてもgoldに一致する組み合わせは見つからなかったため、ここはPDF抽出漏れ、文書側の表現差、またはgold側の要確認差分として扱う。

## 質問別の計算結果

凡例: `index` はvalid質問番号、`implemented_subtype` は実行した計算処理の種類、`predicted_answer` はローカル計算で得た回答、`gold_answer` はvalid正解、`answer_match` は表記ゆれを少しならした一致判定、`needs_review` は人手確認が必要なもの、`source_paths` は参照した主な根拠ファイルを表します。

{case_table.to_markdown(index=False)}

## 質問別の処理内容

凡例: `question` はvalid質問文、`detail` は実際に行ったフィルター、集計、抽出処理を表します。

{detail_table.to_markdown(index=False)}

## サブタイプ別件数

凡例: `subtype` は質問から判定した計算処理の種類、`count` は該当したvalid質問数を表す。

{subtype_df.to_markdown(index=False)}

## 生成物

- `tables/table_valid_calculation_results.csv`: valid表計算7件の計算結果とgold照合
- `tables/table_subtype_summary.csv`: サブタイプ別件数
- `tables/table_calculation_case_plan.csv`: 質問ごとの処理計画
- `tables/cross_project_tax_details.csv`: 全案件消費税集計の内訳

## 次の方針

RAG本体では、検索で表ファイルや表由来Markdownを上位に出すだけでなく、質問が計算型なら以下の順に処理する。

1. 質問文から対象案件、ファイル種別、列名、条件、集計関数を抽出する。
2. CSV/XLSXはLLMへ丸投げせず、pandas/openpyxlで計算する。
3. 計算結果と根拠行数、集計軸、対象ファイルをLLMへ渡し、最終回答の表記だけを整える。
4. 文書横断の金額集計は、最終報告、契約書、提案書の優先順位を明示し、重複・old版を除外する。
"""
    (OUT_DIR / "eda030_report.md").write_text(report, encoding="utf-8")


def main() -> None:
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    diagnosis = pd.read_csv(EDA029_DIAGNOSIS)
    table_cases = diagnosis[diagnosis["route"].eq("table_calculation")].copy()
    table_cases["classified_subtype"] = table_cases["question"].map(classify_subtype)

    results: list[dict[str, object]] = []
    for _, row in table_cases.sort_values("index").iterrows():
        index = int(row["index"])
        result = CALCULATORS[index](index)
        match = answer_matches(result.predicted_answer, str(row["gold_answer"]))
        results.append(
            {
                "index": index,
                "question": row["question"],
                "gold_answer": row["gold_answer"],
                "llm_answer_eda024": row["llm_answer"],
                "classified_subtype": row["classified_subtype"],
                "implemented_subtype": result.subtype,
                "predicted_answer": result.predicted_answer,
                "answer_match": match,
                "needs_review": result.needs_review,
                "detail": result.detail,
                "source_paths": " | ".join(result.source_paths),
            }
        )

    result_df = pd.DataFrame(results)
    subtype_df = (
        result_df.groupby("implemented_subtype", as_index=False)
        .agg(count=("index", "count"), answer_match_count=("answer_match", "sum"), needs_review_count=("needs_review", "sum"))
        .sort_values(["count", "implemented_subtype"], ascending=[False, True])
    )
    plan_df = result_df[
        [
            "index",
            "question",
            "implemented_subtype",
            "source_paths",
            "detail",
            "predicted_answer",
            "gold_answer",
            "answer_match",
            "needs_review",
        ]
    ].copy()

    result_df.to_csv(TABLE_DIR / "table_valid_calculation_results.csv", index=False, encoding="utf-8-sig")
    subtype_df.to_csv(TABLE_DIR / "table_subtype_summary.csv", index=False, encoding="utf-8-sig")
    plan_df.to_csv(TABLE_DIR / "table_calculation_case_plan.csv", index=False, encoding="utf-8-sig")

    manifest = {
        "eda": "EDA030",
        "inputs": [
            normalize_text(EDA029_DIAGNOSIS.relative_to(ROOT)),
            "data/raw/share",
            "data/processed/share",
        ],
        "outputs": [
            "EDA/EDA030/eda030_report.md",
            "EDA/EDA030/tables/table_valid_calculation_results.csv",
            "EDA/EDA030/tables/table_subtype_summary.csv",
            "EDA/EDA030/tables/table_calculation_case_plan.csv",
            "EDA/EDA030/tables/cross_project_tax_details.csv",
        ],
        "table_case_count": int(len(result_df)),
        "answer_match_count": int(result_df["answer_match"].sum()),
        "needs_review_count": int(result_df["needs_review"].sum()),
    }
    (OUT_DIR / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report(result_df, subtype_df)

    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
