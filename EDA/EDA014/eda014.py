from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import re
import shutil
import unicodedata
from datetime import date
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# =============================================================================
# パス設定
# =============================================================================

# eda014.py は「プロジェクト直下 / EDA / EDA014 / eda014.py」に置く前提。
BASE_DIR = Path(__file__).resolve().parents[2]
RAW_SHARE_DIR = BASE_DIR / "data" / "raw" / "share"
RAW_DRIVE_DIR = RAW_SHARE_DIR / "share" / "共有ドライブ"
PROCESSED_SHARE_DIR = BASE_DIR / "data" / "processed" / "share"

OUTPUT_DIR = Path(__file__).resolve().parent
TABLE_DIR = OUTPUT_DIR / "tables"
REPORT_PATH = OUTPUT_DIR / "eda014_report.md"
MANIFEST_PATH = OUTPUT_DIR / "manifest.json"
LOG_PATH = OUTPUT_DIR / "eda014.log"

SUPPORTED_SUFFIXES = {".xlsx", ".csv", ".tsv"}


def setup() -> None:
    """出力フォルダとログ設定を準備する。"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_SHARE_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        filename=LOG_PATH,
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        encoding="utf-8",
        force=True,
    )


def normalize_path_text(text: Any) -> str:
    """Windows上の濁点表記揺れなどをNFCへ寄せる。"""
    return unicodedata.normalize("NFC", str(text))


def relative(path: Path) -> str:
    """レポート用にプロジェクト相対パスを返す。"""
    try:
        return path.resolve().relative_to(BASE_DIR.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def file_sha1(path: Path) -> str:
    """入力ファイルの追跡用にSHA1を計算する。"""
    h = hashlib.sha1()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def safe_filename(text: str) -> str:
    """シート名などをファイル名に使える文字へ寄せる。"""
    value = normalize_path_text(text).strip()
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value)
    return value[:80] or "sheet"


def markdown_escape(text: Any) -> str:
    """Markdown表を壊しやすい文字だけを逃がす。"""
    value = "" if text is None else str(text)
    return value.replace("\\", "\\\\").replace("|", "\\|").replace("\n", " ")


def json_default(value: Any) -> Any:
    """pandasやExcel由来の日付・欠損値をJSON保存できる値へ変換する。"""
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.isoformat()
    if pd.isna(value):
        return None
    return str(value)


def markdown_table(rows: list[dict[str, Any]], max_rows: int = 20) -> str:
    """追加依存なしでMarkdown表を作る。"""
    if not rows:
        return "該当データはありません。"
    columns = list(rows[0].keys())
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join("---" for _ in columns) + " |",
    ]
    for row in rows[:max_rows]:
        lines.append("| " + " | ".join(markdown_escape(row.get(col, "")) for col in columns) + " |")
    return "\n".join(lines)


def save_csv(rows: list[dict[str, Any]], path: Path) -> None:
    """Excelでも読みやすいUTF-8 BOM付きCSVを保存する。"""
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    columns = list(dict.fromkeys(col for row in rows for col in row.keys()))
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def output_paths(raw_path: Path) -> tuple[Path, Path]:
    """raw/shareからの相対構造を保ったMarkdown/JSON出力先を返す。"""
    rel = raw_path.relative_to(RAW_SHARE_DIR)
    output_base = PROCESSED_SHARE_DIR / normalize_path_text(rel.as_posix())
    output_base.parent.mkdir(parents=True, exist_ok=True)
    md_path = output_base.with_suffix(output_base.suffix + ".md")
    json_path = output_base.with_suffix(output_base.suffix + ".structure.json")
    return md_path, json_path


def is_temporary_office_file(path: Path) -> bool:
    """Officeの一時ファイルを除外する。"""
    return path.name.startswith("~$")


def dataframe_profile(df: pd.DataFrame, sample_rows: int) -> dict[str, Any]:
    """DataFrameの列型、欠損、統計、サンプルをJSON化する。"""
    profile: dict[str, Any] = {
        "row_count": int(len(df)),
        "column_count": int(len(df.columns)),
        "columns": [str(col) for col in df.columns],
        "dtypes": {str(col): str(dtype) for col, dtype in df.dtypes.items()},
        "missing_count": {str(col): int(value) for col, value in df.isna().sum().items()},
        "sample_rows": df.head(sample_rows).where(pd.notna(df.head(sample_rows)), None).to_dict(orient="records"),
    }
    numeric = df.select_dtypes(include="number")
    if not numeric.empty:
        profile["numeric_describe"] = (
            numeric.describe()
            .transpose()
            .round(6)
            .where(pd.notna(numeric.describe().transpose()), None)
            .to_dict(orient="index")
        )
    return profile


def read_delimited(path: Path, sep: str) -> tuple[pd.DataFrame, str]:
    """複数エンコーディングを試してCSV/TSVを読む。"""
    last_error = ""
    for encoding in ["utf-8-sig", "utf-8", "cp932"]:
        try:
            return pd.read_csv(path, sep=sep, encoding=encoding), encoding
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)
    raise RuntimeError(last_error)


def process_delimited(path: Path, sample_rows: int) -> tuple[dict[str, Any], str]:
    """CSV/TSVをMarkdown、JSON、正規化CSVへ変換する。"""
    sep = "\t" if path.suffix.lower() == ".tsv" else ","
    df, encoding = read_delimited(path, sep=sep)
    md_path, json_path = output_paths(path)
    normalized_path = md_path.with_suffix(".data.csv")
    normalized_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(normalized_path, index=False, encoding="utf-8-sig")

    profile = dataframe_profile(df, sample_rows=sample_rows)
    record = {
        "raw_relative_path": normalize_path_text(path.relative_to(RAW_SHARE_DIR).as_posix()),
        "processed_markdown_path": relative(md_path),
        "processed_structure_path": relative(json_path),
        "normalized_csv_path": relative(normalized_path),
        "file_name": path.name,
        "file_type": path.suffix.lower().lstrip("."),
        "source_sha1": file_sha1(path),
        "encoding": encoding,
        **profile,
    }
    json_path.write_text(json.dumps(record, ensure_ascii=False, indent=2, default=json_default) + "\n", encoding="utf-8")

    sample_table = markdown_table(profile["sample_rows"], max_rows=min(sample_rows, 10))
    lines = [
        f"# Table Data: {path.name}",
        "",
        "## Source",
        f"- raw_path: `{record['raw_relative_path']}`",
        f"- source_sha1: `{record['source_sha1']}`",
        f"- file_type: `{record['file_type']}`",
        f"- encoding: `{encoding}`",
        f"- row_count: {record['row_count']}",
        f"- column_count: {record['column_count']}",
        f"- normalized_csv_path: `{record['normalized_csv_path']}`",
        "",
        "## Columns",
        "",
        markdown_table(
            [
                {
                    "column": col,
                    "dtype": profile["dtypes"].get(col, ""),
                    "missing_count": profile["missing_count"].get(col, 0),
                }
                for col in profile["columns"]
            ],
            max_rows=80,
        ),
        "",
        "凡例: `column` は列名、`dtype` はpandasで推定した型、`missing_count` は欠損件数を表します。",
        "",
        "## Sample Rows",
        "",
        sample_table,
        "",
        "凡例: 各列は元データの列、各行は先頭サンプル行を表します。",
    ]
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return record, "ok"


def cell_style(cell: Any) -> dict[str, Any]:
    """検索や書式質問に使うセル書式を抽出する。"""
    font = cell.font
    fill = cell.fill
    return {
        "coordinate": cell.coordinate,
        "bold": bool(font.bold),
        "italic": bool(font.italic),
        "underline": bool(font.underline),
        "font_color": str(font.color.rgb) if font.color and font.color.type == "rgb" and font.color.rgb else "",
        "fill_color": str(fill.fgColor.rgb) if fill and fill.fgColor and fill.fgColor.type == "rgb" else "",
        "number_format": cell.number_format or "",
    }


def useful_style(style: dict[str, Any]) -> bool:
    """デフォルトに近い書式はJSONを膨らませないため除外する。"""
    return any(
        [
            style["bold"],
            style["italic"],
            style["underline"],
            bool(style["font_color"]),
            bool(style["fill_color"]) and style["fill_color"] not in {"00000000", "FFFFFFFF"},
            style["number_format"] not in {"General", ""},
        ]
    )


def chart_record(chart: Any, index: int) -> dict[str, Any]:
    """openpyxlのchartオブジェクトから分かる範囲のメタデータを取り出す。"""
    title = ""
    try:
        title = str(chart.title.tx.rich.p[0].r[0].t) if chart.title and chart.title.tx else ""
    except Exception:  # noqa: BLE001
        title = str(chart.title) if chart.title else ""
    return {
        "chart_index": index,
        "chart_type": chart.__class__.__name__,
        "title": title,
        "anchor": str(getattr(chart, "anchor", "")),
        "series_count": len(getattr(chart, "series", []) or []),
    }


def worksheet_to_dataframe(ws: Any) -> pd.DataFrame:
    """ワークシートの値をDataFrameに変換する。先頭行をヘッダー候補として使う。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    max_len = max(len(row) for row in rows)
    values = [list(row) + [None] * (max_len - len(row)) for row in rows]
    header = values[0]
    if all(value is not None and str(value).strip() for value in header):
        columns = [str(value) for value in header]
        data = values[1:]
    else:
        columns = [f"col_{i + 1}" for i in range(max_len)]
        data = values
    return pd.DataFrame(data, columns=columns)


def process_xlsx(path: Path, sample_rows: int, style_limit: int, formula_limit: int) -> tuple[dict[str, Any], str]:
    """Excelブックをシート別CSV、Markdown概要、構造JSONへ変換する。"""
    if is_temporary_office_file(path):
        raise RuntimeError("temporary_office_file")

    wb = load_workbook(path, data_only=False, read_only=False)
    md_path, json_path = output_paths(path)
    sheets_dir = md_path.with_suffix(".sheets")
    if sheets_dir.exists():
        shutil.rmtree(sheets_dir)
    sheets_dir.mkdir(parents=True, exist_ok=True)

    sheet_records: list[dict[str, Any]] = []
    total_formula_count = 0
    total_chart_count = 0
    for ws in wb.worksheets:
        df = worksheet_to_dataframe(ws)
        csv_path = sheets_dir / f"{safe_filename(ws.title)}.csv"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")

        formulas: list[dict[str, Any]] = []
        styles: list[dict[str, Any]] = []
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and len(formulas) < formula_limit:
                    formulas.append({"coordinate": cell.coordinate, "formula": value})
                style = cell_style(cell)
                if useful_style(style) and len(styles) < style_limit:
                    styles.append(style)

        charts = [chart_record(chart, i + 1) for i, chart in enumerate(getattr(ws, "_charts", []) or [])]
        total_formula_count += len(formulas)
        total_chart_count += len(charts)
        profile = dataframe_profile(df, sample_rows=sample_rows)
        sheet_records.append(
            {
                "sheet_name": ws.title,
                "sheet_state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimension": ws.calculate_dimension(),
                "exported_csv_path": relative(csv_path),
                "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
                "auto_filter_ref": ws.auto_filter.ref or "",
                "hidden_rows": [idx for idx, dim in ws.row_dimensions.items() if dim.hidden],
                "hidden_columns": [idx for idx, dim in ws.column_dimensions.items() if dim.hidden],
                "formulas": formulas,
                "formula_count_captured": len(formulas),
                "styled_cells": styles,
                "styled_cell_count_captured": len(styles),
                "charts": charts,
                "chart_count": len(charts),
                "profile": profile,
            }
        )

    record = {
        "raw_relative_path": normalize_path_text(path.relative_to(RAW_SHARE_DIR).as_posix()),
        "processed_markdown_path": relative(md_path),
        "processed_structure_path": relative(json_path),
        "file_name": path.name,
        "file_type": "xlsx",
        "source_sha1": file_sha1(path),
        "sheet_count": len(sheet_records),
        "total_formula_count_captured": total_formula_count,
        "total_chart_count": total_chart_count,
        "sheets": sheet_records,
    }
    json_path.write_text(json.dumps(record, ensure_ascii=False, indent=2, default=json_default) + "\n", encoding="utf-8")

    sheet_summary_rows = [
        {
            "sheet_name": sheet["sheet_name"],
            "rows": sheet["profile"]["row_count"],
            "columns": sheet["profile"]["column_count"],
            "dimension": sheet["dimension"],
            "formulas_captured": sheet["formula_count_captured"],
            "styled_cells_captured": sheet["styled_cell_count_captured"],
            "chart_count": sheet["chart_count"],
            "exported_csv_path": sheet["exported_csv_path"],
        }
        for sheet in sheet_records
    ]
    lines = [
        f"# Excel Workbook: {path.name}",
        "",
        "## Source",
        f"- raw_path: `{record['raw_relative_path']}`",
        f"- source_sha1: `{record['source_sha1']}`",
        f"- sheet_count: {record['sheet_count']}",
        f"- total_chart_count: {record['total_chart_count']}",
        "",
        "## Sheets",
        "",
        markdown_table(sheet_summary_rows, max_rows=80),
        "",
        "凡例: `rows` と `columns` はDataFrame化後の行列数、`dimension` はExcel上の使用範囲、`formulas_captured` は保存した数式セル数、`styled_cells_captured` は保存した書式セル数、`chart_count` はシート内グラフ数を表します。",
    ]
    for sheet in sheet_records[:5]:
        lines.extend(
            [
                "",
                f"## Sample: {sheet['sheet_name']}",
                "",
                markdown_table(sheet["profile"]["sample_rows"], max_rows=min(sample_rows, 10)),
                "",
                "凡例: 各列はシートから推定した列、各行は先頭サンプル行を表します。",
            ]
        )
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return record, "ok"


def process_file(path: Path, args: argparse.Namespace) -> dict[str, Any]:
    """1ファイルを種類に応じて処理し、ログ行を返す。"""
    suffix = path.suffix.lower()
    row = {
        "raw_relative_path": normalize_path_text(path.relative_to(RAW_SHARE_DIR).as_posix()),
        "extension": suffix,
        "status": "",
        "error_type": "",
        "error_message": "",
        "processed_markdown_path": "",
        "processed_structure_path": "",
        "row_count": "",
        "sheet_count": "",
        "chart_count": "",
    }
    try:
        if suffix == ".xlsx":
            record, status = process_xlsx(path, args.sample_rows, args.style_limit, args.formula_limit)
            row.update(
                {
                    "status": status,
                    "processed_markdown_path": record["processed_markdown_path"],
                    "processed_structure_path": record["processed_structure_path"],
                    "sheet_count": record["sheet_count"],
                    "chart_count": record["total_chart_count"],
                }
            )
        elif suffix in {".csv", ".tsv"}:
            record, status = process_delimited(path, args.sample_rows)
            row.update(
                {
                    "status": status,
                    "processed_markdown_path": record["processed_markdown_path"],
                    "processed_structure_path": record["processed_structure_path"],
                    "row_count": record["row_count"],
                }
            )
        else:
            row["status"] = "skipped_unsupported"
    except Exception as exc:  # noqa: BLE001
        row["status"] = "error"
        row["error_type"] = exc.__class__.__name__
        row["error_message"] = str(exc)[:1000]
        logging.exception("failed: %s", path)
    return row


def collect_targets() -> list[Path]:
    """共有ドライブ配下からExcel/CSV/TSVを集める。"""
    return sorted(
        [path for path in RAW_DRIVE_DIR.rglob("*") if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES],
        key=lambda p: normalize_path_text(p.as_posix()).lower(),
    )


def write_report(log_rows: list[dict[str, Any]]) -> None:
    """EDA014の処理結果をMarkdownレポートに保存する。"""
    status_counts = pd.DataFrame(log_rows).groupby(["extension", "status"], dropna=False).size().reset_index(name="count")
    status_rows = status_counts.to_dict(orient="records")
    error_rows = [row for row in log_rows if row["status"] != "ok"]
    lines = [
        "# EDA014: Excel/CSV/TSV表データ前処理",
        "",
        "## 目的",
        "",
        "Excel、CSV、TSVを、LLMが読めるMarkdown、再現・検索用JSON、計算可能なCSVへ変換する。",
        "表データはLLMに丸投げせず、後続のpandas/openpyxl処理で計算できる形を保持する。",
        "",
        "## 出力",
        "",
        "- Markdown/JSON: `data/processed/share/**/*.xlsx.md`, `*.xlsx.structure.json`, `*.csv.md`, `*.csv.structure.json`",
        "- ExcelシートCSV: `data/processed/share/**/*.xlsx.sheets/*.csv`",
        "- 正規化CSV: `data/processed/share/**/*.data.csv`",
        f"- 変換ログ: `{relative(TABLE_DIR / 'tabular_conversion_log.csv')}`",
        "",
        "## 処理結果",
        "",
        "凡例: `extension` は拡張子、`status` は処理状態、`count` は該当ファイル数を表します。",
        "",
        markdown_table(status_rows),
        "",
        "## エラー・スキップ",
        "",
        "凡例: `raw_relative_path` は元ファイル、`extension` は拡張子、`status` は処理状態、`error_type` と `error_message` は失敗理由を表します。",
        "",
        markdown_table(
            [
                {
                    "raw_relative_path": row["raw_relative_path"],
                    "extension": row["extension"],
                    "status": row["status"],
                    "error_type": row["error_type"],
                    "error_message": row["error_message"],
                }
                for row in error_rows
            ],
            max_rows=30,
        ),
        "",
        "## 注意点",
        "",
        "- 暗号化またはOffice一時ファイルは無理に復号せず、エラーまたはスキップとして扱う。",
        "- Excel数式は保存するが、式の再計算はこのEDAでは行わない。",
        "- グラフは存在数と基本メタデータを保存する。画像としての読み取りや数値抽出は別工程で扱う。",
    ]
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_manifest(args: argparse.Namespace, log_rows: list[dict[str, Any]]) -> None:
    """再現用の実行条件を保存する。"""
    manifest = {
        "eda": "EDA014",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "purpose": "Convert Excel/CSV/TSV files into Markdown, JSON metadata, and calculation-ready CSV files.",
        "parameters": {
            "sample_rows": args.sample_rows,
            "style_limit": args.style_limit,
            "formula_limit": args.formula_limit,
        },
        "target_count": len(log_rows),
        "outputs": {
            "report": relative(REPORT_PATH),
            "conversion_log": relative(TABLE_DIR / "tabular_conversion_log.csv"),
            "processed_root": relative(PROCESSED_SHARE_DIR),
        },
        "repro_steps": ["uv run python EDA/EDA014/eda014.py"],
    }
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    """コマンドライン引数を定義する。"""
    parser = argparse.ArgumentParser(description="Convert Excel/CSV/TSV files into preprocessing artifacts.")
    parser.add_argument("--sample-rows", type=int, default=10, help="Markdown/JSONに保存する先頭サンプル行数")
    parser.add_argument("--style-limit", type=int, default=500, help="1シートあたり保存する書式セルの最大数")
    parser.add_argument("--formula-limit", type=int, default=300, help="1シートあたり保存する数式セルの最大数")
    return parser.parse_args()


def main() -> None:
    """EDA014を実行する。"""
    args = parse_args()
    setup()
    targets = collect_targets()
    log_rows = [process_file(path, args) for path in targets]
    save_csv(log_rows, TABLE_DIR / "tabular_conversion_log.csv")
    write_report(log_rows)
    write_manifest(args, log_rows)
    ok_count = sum(1 for row in log_rows if row["status"] == "ok")
    print(f"targets={len(log_rows)} ok={ok_count} errors={len(log_rows) - ok_count}")
    print(f"report={relative(REPORT_PATH)}")


if __name__ == "__main__":
    main()
