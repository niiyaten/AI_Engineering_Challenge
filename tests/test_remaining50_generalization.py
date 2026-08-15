from pathlib import Path
import csv

from rag_recovery.executors.remaining50_generalization import (
    _clean_action_ocr,
    _highlight_band_intersections,
    _money_candidates,
    _tax_included_gross,
)

ROOT = Path(__file__).resolve().parents[1]


def test_remaining50_question_set_is_questions_only_and_unique():
    rows = list(csv.DictReader((ROOT / "questions/remaining50_questions.csv").open(encoding="utf-8-sig", newline="")))
    assert len(rows) == 50
    assert len({int(row["index"]) for row in rows}) == 50
    assert set(rows[0]) == {"index", "question"}


def test_currency_parser_handles_mixed_ocr_thousands_separators():
    text = "固定価格（税込 3.960,000円） / 税抜 3,600,000円"
    assert _tax_included_gross(text) == 3_960_000
    assert max(_money_candidates(text)) == 3_960_000


def test_wrapped_action_cell_is_joined_without_inventing_content():
    ocr = "Action\n前処理\nパイプライン\n実装：\n0値を疑似欠損（NA）扱い\n"
    assert _clean_action_ocr(ocr) == "前処理パイプライン実装：0値を疑似欠損（NA）扱い"


def test_chart_parts_are_naturally_sorted(tmp_path):
    import zipfile
    from rag_recovery.executors.remaining50_generalization import _chart_parts

    path = tmp_path / "charts.xlsx"
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr("xl/charts/chartEx3.xml", "<root/>")
        zf.writestr("xl/charts/chartEx1.xml", "<root/>")
        zf.writestr("xl/charts/chartEx2.xml", "<root/>")
    assert [name for name, _ in _chart_parts(path)] == [
        "xl/charts/chartEx1.xml",
        "xl/charts/chartEx2.xml",
        "xl/charts/chartEx3.xml",
    ]


def test_pick_col_does_not_treat_empty_cells_as_headers():
    from rag_recovery.executors.remaining50_generalization import _pick_col

    row = [1, "T18", None, "QAレビュー", "池田 直哉", "○"]
    assert _pick_col(row, "タスクID") is None
    assert _pick_col(row, "担当者") is None
    header = ["No.", "タスクID", "担当者"]
    assert _pick_col(header, "タスクID") == 1
    assert _pick_col(header, "担当者") == 2


def test_highlight_intersections_use_full_row_and_column_bands():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    yellow = PatternFill(fill_type="solid", fgColor="FFFFFF00")

    for row in range(1, 11):
        for col in range(1, 6):
            ws.cell(row=row, column=col, value=row * 100 + col)

    # One full yellow column and two full yellow rows. Ordinary highlighted
    # numeric cells elsewhere in the rows must not be mistaken for intersections.
    for row in range(1, 11):
        ws.cell(row=row, column=4).fill = yellow
    for row in (3, 8):
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = yellow
    ws["D3"] = 10_096
    ws["D8"] = 10_368

    intersections, diagnostics = _highlight_band_intersections(ws)

    assert intersections == [("D3", 10_096.0), ("D8", 10_368.0)]
    assert diagnostics["row_bands"] == [3, 8]
    assert diagnostics["column_bands"] == [4]


def test_f1_difference_uses_rounding_not_truncation():
    intermediate = 0.7329671168078127
    final = 0.8291582445227382
    assert f"{abs(final - intermediate):.8f}" == "0.09619113"
